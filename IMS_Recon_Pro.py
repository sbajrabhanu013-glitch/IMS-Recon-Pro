import json
import re
import sqlite3
import pickle
from dataclasses import dataclass
from datetime import datetime, date
from io import BytesIO
from copy import copy, deepcopy
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# =========================================================
# IMS RECON PRO — FULL WORKING VERSION
# Premium UI + Login + Purchase Register + IMS JSON + Reco + Action + Final GST JSON
# Copyright @BAJRABHANU
# =========================================================

APP_TITLE = "IMS Recon Pro"
APP_TAGLINE = "Intelligent GST IMS Reconciliation & Action Management Platform"
COPYRIGHT_OWNER = "@BAJRABHANU"
APP_DB = "ims_recon_pro.db"
ENGINE_VERSION = "2026.05.04-V7"

IMS_SHEETS = ["B2B", "B2BA", "B2B-DN", "B2B-DNA", "B2B-CN", "B2B-CNA"]
ACTION_VALUES = ["No Action", "Accepted", "Rejected", "Pending", "Review"]
USER_MASTER = {
    "MainAdmin": {"password": "Adminpwd", "role": "Main Admin", "name": "Main Admin"},
    "User1": {"password": "Userpwd1", "role": "Sub User", "name": "User One"},
    "User2": {"password": "Userpwd2", "role": "Sub User", "name": "User Two"},
}

MONEY_COLS = ["invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess"]
TAX_COLS = ["igst", "cgst", "sgst", "cess"]

COLUMN_ALIASES = {
    "supplier_gstin": [
        "supplier gstin", "gstin of supplier", "gstin", "ctin", "stin", "counterparty gstin",
        "vendor gstin", "party gstin", "gstin/uın of supplier", "gstin/uin of supplier"
    ],
    "supplier_name": [
        "supplier name", "trade/legal name", "trade legal name", "tradenm", "legal name",
        "vendor name", "party name", "name", "supplier legal name"
    ],
    "document_type": [
        "document type", "doc type", "invoice type", "type", "supply type",
        "nature of document", "note type"
    ],
    "document_no": [
        "document number", "document no", "doc no", "invoice number", "invoice no", "inum", "nt_num",
        "invoice", "note number", "note no", "bill number", "voucher number"
    ],
    "document_date": [
        "document date", "doc date", "invoice date", "idt", "nt_dt", "date", "note date", "bill date"
    ],
    "invoice_value": [
        "invoice value", "document value", "val", "total invoice value", "gross value",
        "total value", "invoice value(inr)", "invoice value(rs)", "total document value"
    ],
    "taxable_value": [
        "taxable value", "txval", "taxable amount", "taxable value(inr)", "taxable value(rs)",
        "assessable value", "net value", "taxable val"
    ],
    "igst": ["igst", "iamt", "integrated tax", "integrated tax amount", "igst amount"],
    "cgst": ["cgst", "camt", "central tax", "central tax amount", "cgst amount"],
    "sgst": ["sgst", "samt", "state tax", "state/ut tax", "utgst", "sgst amount"],
    "cess": ["cess", "cess amount"],
    "itc_available": ["itc available", "itc availability", "eligible itc", "itc eligibility", "eligible"],
    "ims_status": ["status", "ims status", "recipient status", "recipient action", "action"],
    "remarks": ["remarks", "remark", "reason", "comments", "comment"],
    "pos": ["place of supply", "pos", "state", "supply state"],
    "return_period": ["return period", "rtnprd", "tax period", "period", "month"],
}


# =========================================================
# STREAMLIT PAGE
# =========================================================

st.set_page_config(
    page_title=f"{APP_TITLE} | {COPYRIGHT_OWNER}",
    page_icon="🇮🇳",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items=None,
)


# =========================================================
# DATABASE
# =========================================================

def get_conn():
    return sqlite3.connect(APP_DB, check_same_thread=False)


def init_db():
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS user_data (
                username TEXT NOT NULL,
                key TEXT NOT NULL,
                value BLOB NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (username, key)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT,
                event_time TEXT,
                event_type TEXT,
                detail TEXT
            )
        """)


def db_save(username: str, key: str, value):
    blob = pickle.dumps(value)
    with get_conn() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO user_data (username, key, value, updated_at) VALUES (?, ?, ?, ?)",
            (username, key, blob, datetime.now().isoformat(timespec="seconds")),
        )


def db_load(username: str, key: str, default=None):
    try:
        with get_conn() as conn:
            row = conn.execute("SELECT value FROM user_data WHERE username=? AND key=?", (username, key)).fetchone()
        if not row:
            return default
        return pickle.loads(row[0])
    except Exception:
        return default


def db_delete_user(username: str):
    with get_conn() as conn:
        conn.execute("DELETE FROM user_data WHERE username=?", (username,))
        conn.execute("DELETE FROM audit_log WHERE username=?", (username,))


def log_event(event_type: str, detail: str):
    username = st.session_state.get("username", "")
    try:
        with get_conn() as conn:
            conn.execute(
                "INSERT INTO audit_log (username, event_time, event_type, detail) VALUES (?, ?, ?, ?)",
                (username, datetime.now().isoformat(timespec="seconds"), event_type, detail),
            )
    except Exception:
        pass


def load_audit(username: str = "") -> pd.DataFrame:
    try:
        with get_conn() as conn:
            if username:
                return pd.read_sql(
                    "SELECT event_time, event_type, detail FROM audit_log WHERE username=? ORDER BY id DESC",
                    conn,
                    params=(username,),
                )
            return pd.read_sql(
                "SELECT username, event_time, event_type, detail FROM audit_log ORDER BY id DESC",
                conn,
            )
    except Exception:
        return pd.DataFrame()


# =========================================================
# SESSION
# =========================================================

def init_state():
    defaults = {
        "logged_in": False,
        "username": "",
        "role": "",
        "display_name": "",
        "page": "Dashboard",
        "client_name": "",
        "client_gstin": "",
        "return_period": datetime.today().strftime("%b-%Y"),
        "purchase_df": pd.DataFrame(),
        "ims_df": pd.DataFrame(),
        "ims_source": "",
        "ims_json_records": [],
        "ims_template_bytes": b"",
        "ims_auto_xlsm_bytes": b"",
        "ims_json_data": {},
        "ims_json_bytes": b"",
        "final_action_xlsm_bytes": b"",
        "final_json_bytes": b"",
        "final_json_summary": pd.DataFrame(),
        "recon_df": pd.DataFrame(),
        "action_df": pd.DataFrame(),
        "amount_tolerance": 5.0,
        "date_tolerance": 2,
        "include_amendments": True,
        "use_fuzzy": False,
    }
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)


def load_user_state():
    username = st.session_state.get("username")
    if not username:
        return
    for key in [
        "client_name", "client_gstin", "return_period",
        "purchase_df", "ims_df", "ims_source", "ims_json_records", "ims_template_bytes", "ims_auto_xlsm_bytes", "ims_json_data", "ims_json_bytes", "final_action_xlsm_bytes", "final_json_bytes", "final_json_summary", "recon_df", "action_df"
    ]:
        st.session_state[key] = db_load(username, key, st.session_state.get(key))


def save_user_state(keys: Optional[List[str]] = None):
    username = st.session_state.get("username")
    if not username:
        return
    keys = keys or [
        "client_name", "client_gstin", "return_period",
        "purchase_df", "ims_df", "ims_source", "ims_json_records", "ims_template_bytes", "ims_auto_xlsm_bytes", "ims_json_data", "ims_json_bytes", "final_action_xlsm_bytes", "final_json_bytes", "final_json_summary", "recon_df", "action_df"
    ]
    for key in keys:
        db_save(username, key, st.session_state.get(key))


# =========================================================
# STYLING
# =========================================================

def inject_css():
    st.markdown("""
    <style>
        :root {
            --navy:#061a3e; --navy2:#0b2d66; --blue:#2563eb; --cyan:#38bdf8;
            --saffron:#ff9933; --green:#138808; --gold:#f3b34d; --red:#dc463f;
            --bg1:#d8e7f7; --bg2:#b9cde8; --card:#ffffff; --border:#c8d8ec;
            --text:#102244; --muted:#5c708f;
        }

        .stApp {
            background:
                radial-gradient(circle at 12% 8%, rgba(255,153,51,0.20), transparent 26%),
                radial-gradient(circle at 88% 16%, rgba(56,189,248,0.18), transparent 28%),
                radial-gradient(circle at 88% 82%, rgba(19,136,8,0.13), transparent 30%),
                linear-gradient(135deg, #dbeafe 0%, #c7d9ef 48%, #e4eef9 100%);
            color: var(--text);
        }

        header[data-testid="stHeader"] {
            background: rgba(224, 236, 249, 0.88) !important;
            backdrop-filter: blur(12px);
            border-bottom: 1px solid rgba(11,45,102,0.10);
        }

        div[data-testid="stToolbar"] { visibility:hidden; height:0; }
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}

        .block-container {
            padding-top: 1.05rem;
            padding-bottom: 2rem;
            max-width: 1560px;
        }

        .gst-shell {
            border-radius: 28px;
            overflow: hidden;
            border: 1px solid rgba(198,216,236,0.95);
            background: #ffffff;
            box-shadow: 0 18px 45px rgba(7, 26, 61, 0.16);
            margin-bottom: 18px;
        }

        .gst-top-strip {
            min-height: 34px;
            background: linear-gradient(90deg, #061a3e 0%, #0b2d66 60%, #061a3e 100%);
            display:flex; align-items:center; justify-content:flex-end;
            gap:18px; padding: 0 22px;
            color:#dceaff; font-size: 12px; letter-spacing:.02em;
        }

        .gst-masthead {
            position:relative;
            background:
                linear-gradient(120deg, rgba(255,153,51,0.14), transparent 28%),
                linear-gradient(90deg, #08214f 0%, #0b3677 52%, #123d82 100%);
            min-height: 150px;
            padding: 28px 32px;
            display:flex; align-items:center; justify-content:space-between;
            gap:24px; overflow:hidden;
        }

        .gst-masthead::after {
            content:""; position:absolute; right:-130px; top:-170px;
            width:420px; height:420px; border-radius:50%;
            background: radial-gradient(circle, rgba(255,255,255,0.16), transparent 62%);
        }

        .gst-brand { position:relative; z-index:2; display:flex; align-items:center; gap:20px; min-width:0; }

        .gst-emblem {
            width:78px; height:78px; border-radius:22px;
            display:flex; align-items:center; justify-content:center;
            color:#ffffff; font-size:42px;
            background: linear-gradient(135deg, rgba(255,153,51,0.95), rgba(19,136,8,0.88));
            border:1px solid rgba(255,255,255,0.22);
            box-shadow:0 14px 30px rgba(0,0,0,0.20);
        }

        .gst-title {
            font-size: 36px; font-weight: 900; letter-spacing: -.02em;
            line-height: 1.05; color:#ffffff;
        }

        .gst-subtitle {
            margin-top: 8px; font-size: 18px; line-height: 1.3;
            color:#edf5ff; font-weight: 500;
        }

        .header-note { margin-top: 7px; font-size: 13px; color:#cfe0fb; }

        .gst-action-wrap {
            position:relative; z-index:2; display:flex; align-items:center;
            justify-content:flex-end; min-width: 220px;
        }

        .gst-floating-flag {
            width: 150px; height: 92px; border-radius: 14px;
            position: relative; overflow: hidden;
            background: linear-gradient(to bottom, #ff9933 0 33.33%, #ffffff 33.33% 66.66%, #138808 66.66% 100%);
            border:1px solid rgba(255,255,255,0.45);
            box-shadow: 0 14px 28px rgba(0,0,0,0.24);
            animation: gstFlagFloat 3.8s ease-in-out infinite;
        }

        .gst-floating-flag::after {
            content:"☸"; position:absolute; left:50%; top:50%;
            transform:translate(-50%, -50%);
            color:#0a3d91; font-size:28px; font-weight:800; z-index:2;
        }

        @keyframes gstFlagFloat {
            0%,100% { transform: translateY(0px); }
            50% { transform: translateY(-5px); }
        }

        .gst-meta-row {
            background: linear-gradient(90deg, #eef5ff 0%, #e7f0fb 100%);
            border-top: 1px solid #c9d9ee;
            padding: 16px 18px;
            display:grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap:14px;
        }

        .gst-meta-card {
            background:#ffffff; border:1px solid #ceddf0; border-radius:18px;
            min-height:72px; padding:14px 16px;
            display:flex; align-items:center; gap:12px;
            box-shadow:0 8px 18px rgba(7,26,61,0.06);
        }

        .gst-meta-icon {
            width:44px; height:44px; border-radius:14px;
            background:linear-gradient(135deg,#eef6ff,#dbeafe);
            display:flex; align-items:center; justify-content:center;
            font-size:21px; flex-shrink:0;
        }

        .gst-meta-label {
            font-size:12px; color:var(--muted); font-weight:700;
            text-transform:uppercase; letter-spacing:.04em;
        }

        .gst-meta-value {
            font-size:15px; color:var(--text); font-weight:800;
            line-height:1.25; margin-top:3px;
        }

        .main-shell, .panel, .metric-card, .small-card {
            background: rgba(255,255,255,0.96);
            border: 1px solid rgba(200,216,236,0.98);
            box-shadow: 0 14px 35px rgba(7,26,61,0.12);
        }

        .main-shell { border-radius: 26px; overflow: hidden; margin-bottom: 18px; }
        .content-pad { padding: 30px; position:relative; }
        .panel, .metric-card, .small-card { border-radius: 24px; padding: 20px 22px; height: 100%; }
        .panel { background: linear-gradient(180deg,#ffffff,#f8fbff); }

        .metric-card { position:relative; overflow:hidden; }
        .metric-card::after {
            content:""; position:absolute; right:-36px; bottom:-48px;
            width:115px; height:115px; border-radius:50%;
            background: rgba(37,99,235,0.07);
        }

        .metric-top {display:flex;align-items:center;gap:15px;position:relative;z-index:2;}
        .metric-icon {
            width:58px;height:58px;border-radius:18px;
            display:flex;align-items:center;justify-content:center;
            font-size:25px;
            box-shadow: inset 0 0 0 1px rgba(255,255,255,0.55);
        }
        .metric-label {
            font-size:13px;color:#60748f;font-weight:800;
            text-transform:uppercase;letter-spacing:.04em;
        }
        .metric-value {font-size:33px;font-weight:900;color:#112244;line-height:1.15;}
        .metric-delta {font-size:13px;color:#12a150;margin-top:5px;font-weight:700;}
        .metric-delta.red {color:#e1563a;}

        .panel-title {font-size:20px;font-weight:900;color:#112244;}
        .section-title {font-size:28px;font-weight:950;color:#102244;margin:10px 0 4px 0;}
        .section-sub {font-size:15px;color:#60748f;margin-bottom:18px;}

        .headline {font-size:21px;color:#e98012;font-weight:900;}
        .main-title {font-size:34px;font-weight:950;color:#112244;line-height:1.18;margin-top:8px;}
        .subcopy {font-size:17px;color:#52637d;margin-top:12px;line-height:1.5;}

        .cta-dark,.cta-light {
            display:inline-block; padding:13px 24px; border-radius:16px;
            font-weight:900; text-decoration:none; font-size:15px;
            margin-right:10px; margin-top:20px;
        }
        .cta-dark {background:linear-gradient(135deg,#0b2d66,#2563eb);color:white;box-shadow:0 12px 22px rgba(11,42,93,.22);}
        .cta-light {background:white;color:#0b2a5d;border:1px solid #d0def0;}

        .feature-card {
            background:linear-gradient(180deg,#fffaf1,#fff7eb);
            border:1px solid #f0dfc0; border-radius:18px;
            padding:15px 17px; margin-bottom:13px;
            box-shadow:0 8px 18px rgba(7,26,61,0.06);
        }
        .feature-card.blue {background:linear-gradient(180deg,#f5f9ff,#eef5ff);border-color:#d6e4ff;}
        .feature-card.green {background:linear-gradient(180deg,#f5fbf3,#eff9ec);border-color:#d8ead0;}
        .feature-title {font-weight:900;color:#23385d;font-size:16px;}
        .feature-desc {font-size:13px;color:#5f6f89;line-height:1.4;margin-top:4px;}
        .shield-center {
            width:140px;height:140px;border-radius:50%;margin:0 auto 18px auto;
            background:radial-gradient(circle at 30% 30%,#fffef4,#f8f0d2 55%,#ead39f 100%);
            display:flex;align-items:center;justify-content:center;font-size:62px;
            box-shadow:inset 0 0 0 10px rgba(255,255,255,.65),0 12px 28px rgba(194,165,97,.18);
        }

        .watermark {
            position:absolute;left:50%;top:47%;transform:translate(-50%,-50%);
            font-size:230px;color:rgba(14,41,90,.035);pointer-events:none;
        }

        .stTextInput input, .stNumberInput input, .stDateInput input, .stTextArea textarea {
            border-radius: 14px !important;
            border: 1px solid #cdddf0 !important;
            background: #ffffff !important;
        }

        .stSelectbox div[data-baseweb="select"] > div {
            border-radius: 14px !important;
            border-color: #cdddf0 !important;
        }

        .stTabs [data-baseweb="tab-list"] {
            gap: 8px; background:#edf4ff; padding:8px;
            border-radius:18px; border:1px solid #cfdded;
        }
        .stTabs [data-baseweb="tab"] {
            border-radius:14px; padding:10px 16px; font-weight:800;
        }
        .stTabs [aria-selected="true"] {
            background:#ffffff !important; color:#0b2d66 !important;
            box-shadow:0 8px 16px rgba(7,26,61,0.08);
        }

        div[data-testid="stDataFrame"] {
            border-radius: 18px; overflow:hidden;
            border:1px solid #cfdff0;
            box-shadow:0 8px 18px rgba(7,26,61,0.06);
        }

        div[data-testid="stHorizontalBlock"] .stButton > button {
            border-radius: 16px !important;
            min-height: 48px !important;
            border: 1px solid #c9d9ee !important;
            background: linear-gradient(180deg, #ffffff, #f2f7ff) !important;
            color: #0d2d63 !important;
            font-weight: 800 !important;
            box-shadow: 0 6px 14px rgba(7,26,61,0.06) !important;
        }

        div[data-testid="stHorizontalBlock"] .stButton > button:hover {
            border-color: #2563eb !important;
            background: linear-gradient(180deg, #eef6ff, #dbeafe) !important;
            color: #071a3d !important;
            transform: translateY(-1px);
        }

        .footer-bar {
            margin-top: 20px; border-radius: 24px;
            background: linear-gradient(90deg,#061a3e 0%,#082b61 50%,#061a3e 100%);
            color:white; padding: 18px 22px;
            box-shadow: 0 14px 32px rgba(7,26,61,0.16);
        }
        .foot-item {display:flex;align-items:center;gap:10px;justify-content:center;}
        .foot-main {font-weight:900;}
        .foot-sub {font-size:13px;color:#d4e0ff;}

        .login-bg {
            min-height:calc(100vh - 35px);
            display:flex; align-items:center; justify-content:center;
            background:
                radial-gradient(circle at 18% 12%, rgba(255,153,51,.20), transparent 30%),
                radial-gradient(circle at 85% 80%, rgba(19,136,8,.16), transparent 32%),
                linear-gradient(135deg,#071a3d,#0d2d63);
            border-radius:30px; position:relative; overflow:hidden;
        }
        .login-bg::after {
            content:""; position:absolute;
            width:720px;height:720px;border-radius:50%;
            right:-220px;top:-260px;
            background:radial-gradient(circle, rgba(255,255,255,0.13), transparent 62%);
        }
        .login-card {
            width: 470px; background:rgba(255,255,255,.96);
            border:1px solid rgba(255,255,255,.62);
            border-radius:32px; padding:38px;
            box-shadow:0 35px 90px rgba(0,0,0,.30);
            position:relative; z-index:2;
        }
        .login-title {font-size:36px;font-weight:950;color:#071b4a;text-align:center;}
        .login-sub {font-size:15px;color:#566982;text-align:center;margin-bottom:24px;line-height:1.5;}

        .copyright-float {
            position:fixed;right:18px;bottom:14px;
            color:rgba(7,26,61,.20);
            font-weight:950;letter-spacing:.08em;z-index:99;
        }

        @media (max-width: 1100px) {
            .gst-masthead {flex-direction:column;align-items:flex-start;}
            .gst-meta-row {grid-template-columns: repeat(2, minmax(0, 1fr));}
            .gst-action-wrap {justify-content:flex-start;}
        }
        @media (max-width: 760px) {
            .gst-title {font-size:28px;}
            .gst-subtitle {font-size:15px;}
            .gst-meta-row {grid-template-columns: 1fr;}
            .block-container {padding-left:0.75rem;padding-right:0.75rem;}
        }
    
        /* ================= V9 SALEABLE UI EDITION ================= */
        .v9-workflow {
            background: rgba(255,255,255,0.96);
            border: 1px solid #c9d9ee;
            border-radius: 24px;
            padding: 16px 18px;
            margin: 0 0 18px 0;
            box-shadow: 0 14px 34px rgba(7,26,61,0.10);
        }
        .v9-workflow-title {
            font-size: 17px;
            font-weight: 950;
            color: #0b2d66;
            margin-bottom: 12px;
            display:flex;
            align-items:center;
            gap:8px;
        }
        .v9-step-grid {
            display:grid;
            grid-template-columns: repeat(6, minmax(0, 1fr));
            gap: 10px;
        }
        .v9-step {
            position:relative;
            min-height: 88px;
            border-radius: 18px;
            border: 1px solid #d1e0f2;
            background: linear-gradient(180deg,#ffffff,#f4f8ff);
            padding: 12px;
            overflow:hidden;
        }
        .v9-step.done {
            border-color: rgba(19,136,8,0.28);
            background: linear-gradient(180deg,#ffffff,#effaf1);
        }
        .v9-step.active {
            border-color: rgba(255,153,51,0.55);
            background: linear-gradient(180deg,#ffffff,#fff5e8);
            box-shadow: inset 0 0 0 1px rgba(255,153,51,0.16);
        }
        .v9-step.pending {
            border-color: rgba(120,140,170,0.25);
        }
        .v9-step-num {
            width:30px;height:30px;border-radius:10px;
            display:flex;align-items:center;justify-content:center;
            font-weight:900;font-size:13px;
            background:#eaf2ff;color:#0b2d66;margin-bottom:8px;
        }
        .v9-step.done .v9-step-num { background:#e9f9ed;color:#138808; }
        .v9-step.active .v9-step-num { background:#fff0da;color:#c76f00; }
        .v9-step-label {font-size:13px;font-weight:900;color:#102244;line-height:1.25;}
        .v9-step-status {font-size:11px;font-weight:800;color:#5c708f;margin-top:5px;}
        .v9-step.done .v9-step-status {color:#138808;}
        .v9-step.active .v9-step-status {color:#c76f00;}

        .v9-kpi-strip {
            display:grid;
            grid-template-columns: repeat(4, minmax(0,1fr));
            gap:14px;
            margin: 14px 0 18px 0;
        }
        .v9-kpi {
            background:linear-gradient(135deg,#ffffff,#f7fbff);
            border:1px solid #cdddf0;
            border-radius:22px;
            padding:18px;
            box-shadow:0 12px 26px rgba(7,26,61,0.09);
            min-height:112px;
            position:relative;
            overflow:hidden;
        }
        .v9-kpi::after {
            content:"";
            position:absolute;
            width:105px;height:105px;right:-38px;bottom:-48px;
            border-radius:50%;background:rgba(37,99,235,0.08);
        }
        .v9-kpi-label {font-size:12px;text-transform:uppercase;letter-spacing:.04em;font-weight:900;color:#60748f;}
        .v9-kpi-value {font-size:32px;font-weight:950;color:#102244;margin-top:8px;}
        .v9-kpi-note {font-size:12px;font-weight:800;color:#138808;margin-top:6px;}

        .v9-module-grid {
            display:grid;
            grid-template-columns: repeat(3, minmax(0,1fr));
            gap:16px;
            margin: 14px 0 20px 0;
        }
        .v9-module-card {
            background:#ffffff;
            border:1px solid #cdddf0;
            border-radius:24px;
            padding:20px;
            min-height:185px;
            box-shadow:0 14px 34px rgba(7,26,61,0.10);
            position:relative;
            overflow:hidden;
        }
        .v9-module-card::before {
            content:"";
            position:absolute;left:0;top:0;width:100%;height:6px;
            background:linear-gradient(90deg,#ff9933,#2563eb,#138808);
        }
        .v9-module-icon {
            width:52px;height:52px;border-radius:16px;
            display:flex;align-items:center;justify-content:center;
            background:linear-gradient(135deg,#eef6ff,#dbeafe);
            font-size:25px;margin-bottom:12px;
        }
        .v9-module-title {font-size:18px;font-weight:950;color:#102244;margin-bottom:7px;}
        .v9-module-desc {font-size:13px;color:#60748f;line-height:1.45;}
        .v9-module-badge {
            display:inline-block;margin-top:12px;padding:6px 11px;border-radius:999px;
            font-size:11px;font-weight:900;background:#eef6ff;color:#0b2d66;border:1px solid #d1e0f2;
        }

        .v9-readiness {
            background:linear-gradient(135deg,#071a3d,#0b3677);
            color:#ffffff;border-radius:26px;padding:24px;
            border:1px solid rgba(255,255,255,0.18);
            box-shadow:0 18px 40px rgba(7,26,61,0.18);
            margin: 14px 0 18px 0;
        }
        .v9-readiness-title {font-size:22px;font-weight:950;margin-bottom:12px;}
        .v9-check-grid {
            display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;
        }
        .v9-check {
            background:rgba(255,255,255,0.10);
            border:1px solid rgba(255,255,255,0.14);
            border-radius:18px;padding:14px;
            min-height:82px;
        }
        .v9-check-icon {font-size:22px;margin-bottom:6px;}
        .v9-check-label {font-size:13px;font-weight:900;color:#eef5ff;line-height:1.3;}

        .v9-action-badge {
            display:inline-block;padding:5px 10px;border-radius:999px;
            font-size:12px;font-weight:900;border:1px solid transparent;
        }
        .v9-action-accepted {background:#e9f9ed;color:#138808;border-color:#bfe9c9;}
        .v9-action-pending {background:#fff3df;color:#b96b00;border-color:#ffd9a8;}
        .v9-action-rejected {background:#fff0ed;color:#d33a2f;border-color:#ffc8c0;}
        .v9-action-review {background:#f2ecff;color:#6d3bd1;border-color:#d8c9ff;}
        .v9-action-no {background:#f1f5f9;color:#475569;border-color:#d7e0ea;}

        .v9-help-box {
            background:#fffdf6;border:1px solid #f3d9a7;border-radius:20px;
            padding:16px 18px;margin:12px 0 18px 0;
            color:#62420d;box-shadow:0 10px 22px rgba(7,26,61,0.06);
        }
        .v9-help-title {font-weight:950;font-size:16px;margin-bottom:6px;color:#7a4b00;}
        .v9-help-text {font-size:13px;line-height:1.5;}

        .v9-report-grid {
            display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:16px;
            margin:16px 0;
        }
        .v9-report-card {
            background:#ffffff;border:1px solid #cdddf0;border-radius:22px;
            padding:18px;box-shadow:0 12px 26px rgba(7,26,61,0.09);
        }
        .v9-report-title {font-size:16px;font-weight:950;color:#102244;}
        .v9-report-desc {font-size:13px;color:#60748f;margin-top:7px;line-height:1.45;}

        @media (max-width: 1150px) {
            .v9-step-grid {grid-template-columns: repeat(3, minmax(0,1fr));}
            .v9-module-grid, .v9-report-grid {grid-template-columns: repeat(2,minmax(0,1fr));}
            .v9-kpi-strip, .v9-check-grid {grid-template-columns: repeat(2,minmax(0,1fr));}
        }
        @media (max-width: 760px) {
            .v9-step-grid, .v9-module-grid, .v9-report-grid, .v9-kpi-strip, .v9-check-grid {
                grid-template-columns: 1fr;
            }
        }

    
        /* ================= V10 ADVANCED SALEABLE UI ================= */
        .v10-command-center {
            background: linear-gradient(135deg,#071a3d,#0b3677);
            color:#ffffff;
            border-radius:28px;
            padding:24px;
            box-shadow:0 20px 48px rgba(7,26,61,0.22);
            border:1px solid rgba(255,255,255,0.16);
            margin: 14px 0 20px 0;
            position:relative;
            overflow:hidden;
        }
        .v10-command-center::after {
            content:"";
            position:absolute;
            right:-120px;
            top:-160px;
            width:360px;
            height:360px;
            border-radius:50%;
            background:radial-gradient(circle,rgba(255,255,255,0.15),transparent 62%);
        }
        .v10-command-title {
            font-size:26px;
            font-weight:950;
            margin-bottom:8px;
            position:relative;
            z-index:2;
        }
        .v10-command-sub {
            font-size:14px;
            color:#d9e8ff;
            margin-bottom:18px;
            position:relative;
            z-index:2;
        }
        .v10-action-grid {
            display:grid;
            grid-template-columns:repeat(4,minmax(0,1fr));
            gap:14px;
            position:relative;
            z-index:2;
        }
        .v10-action-card {
            background:rgba(255,255,255,0.10);
            border:1px solid rgba(255,255,255,0.16);
            border-radius:22px;
            padding:18px;
            min-height:126px;
            backdrop-filter: blur(4px);
        }
        .v10-action-icon {font-size:30px;margin-bottom:10px;}
        .v10-action-title {font-size:16px;font-weight:950;color:#fff;}
        .v10-action-desc {font-size:12px;color:#d7e7ff;line-height:1.45;margin-top:6px;}

        .v10-quality-grid {
            display:grid;
            grid-template-columns:repeat(2,minmax(0,1fr));
            gap:18px;
            margin:16px 0 18px 0;
        }
        .v10-quality-card {
            background:#ffffff;
            border:1px solid #cdddf0;
            border-radius:26px;
            padding:22px;
            box-shadow:0 14px 34px rgba(7,26,61,0.10);
        }
        .v10-quality-head {
            display:flex;
            justify-content:space-between;
            align-items:flex-start;
            gap:16px;
            margin-bottom:16px;
        }
        .v10-quality-title {
            font-size:20px;
            font-weight:950;
            color:#102244;
        }
        .v10-quality-score {
            min-width:86px;
            height:86px;
            border-radius:50%;
            display:flex;
            align-items:center;
            justify-content:center;
            font-size:24px;
            font-weight:950;
            color:#ffffff;
            background:linear-gradient(135deg,#138808,#28b463);
            box-shadow:0 12px 26px rgba(19,136,8,0.22);
        }
        .v10-quality-score.warn {background:linear-gradient(135deg,#ff9933,#f3b34d);}
        .v10-quality-score.bad {background:linear-gradient(135deg,#dc463f,#ef675b);}
        .v10-mini-grid {
            display:grid;
            grid-template-columns:repeat(3,minmax(0,1fr));
            gap:10px;
        }
        .v10-mini-stat {
            background:#f7fbff;
            border:1px solid #d6e4f5;
            border-radius:16px;
            padding:12px;
        }
        .v10-mini-label {font-size:11px;color:#60748f;font-weight:900;text-transform:uppercase;}
        .v10-mini-value {font-size:18px;color:#102244;font-weight:950;margin-top:5px;}

        .v10-control-room {
            display:grid;
            grid-template-columns:1.35fr .9fr;
            gap:18px;
            margin:16px 0 20px 0;
        }
        .v10-control-main,
        .v10-control-side {
            background:#ffffff;
            border:1px solid #cdddf0;
            border-radius:26px;
            padding:22px;
            box-shadow:0 14px 34px rgba(7,26,61,0.10);
        }
        .v10-control-title {
            font-size:22px;
            font-weight:950;
            color:#102244;
            margin-bottom:12px;
        }
        .v10-badge-row {
            display:flex;
            gap:10px;
            flex-wrap:wrap;
            margin:10px 0 14px 0;
        }
        .v10-filter-badge {
            display:inline-flex;
            align-items:center;
            gap:6px;
            padding:8px 12px;
            border-radius:999px;
            background:#eef6ff;
            border:1px solid #d1e0f2;
            color:#0b2d66;
            font-size:12px;
            font-weight:900;
        }
        .v10-filter-badge.green {background:#e9f9ed;color:#138808;border-color:#bfe9c9;}
        .v10-filter-badge.orange {background:#fff3df;color:#b96b00;border-color:#ffd9a8;}
        .v10-filter-badge.red {background:#fff0ed;color:#d33a2f;border-color:#ffc8c0;}
        .v10-filter-badge.purple {background:#f2ecff;color:#6d3bd1;border-color:#d8c9ff;}

        .v10-empty-state {
            background:linear-gradient(135deg,#ffffff,#f7fbff);
            border:1px dashed #a8bdd8;
            border-radius:26px;
            padding:28px;
            text-align:center;
            margin:16px 0;
            box-shadow:0 12px 28px rgba(7,26,61,0.07);
        }
        .v10-empty-icon {font-size:44px;margin-bottom:10px;}
        .v10-empty-title {font-size:22px;font-weight:950;color:#102244;}
        .v10-empty-text {font-size:14px;color:#60748f;margin-top:8px;}

        .v10-management-summary {
            background:linear-gradient(135deg,#ffffff,#f6faff);
            border:1px solid #cdddf0;
            border-radius:26px;
            padding:22px;
            box-shadow:0 14px 34px rgba(7,26,61,0.10);
            margin:16px 0 18px 0;
        }
        .v10-management-title {
            font-size:22px;
            font-weight:950;
            color:#102244;
            margin-bottom:10px;
        }
        .v10-management-text {
            font-size:15px;
            line-height:1.6;
            color:#334866;
        }

        .v10-json-review {
            background:#ffffff;
            border:1px solid #cdddf0;
            border-radius:28px;
            padding:24px;
            box-shadow:0 16px 38px rgba(7,26,61,0.12);
            margin:16px 0 18px 0;
        }
        .v10-json-title {
            font-size:24px;
            font-weight:950;
            color:#102244;
            margin-bottom:14px;
        }
        .v10-json-checks {
            display:grid;
            grid-template-columns:repeat(5,minmax(0,1fr));
            gap:12px;
        }
        .v10-json-check {
            background:#f7fbff;
            border:1px solid #d6e4f5;
            border-radius:18px;
            padding:14px;
            min-height:96px;
            text-align:center;
        }
        .v10-json-check-icon {font-size:24px;margin-bottom:7px;}
        .v10-json-check-label {font-size:12px;font-weight:900;color:#102244;line-height:1.35;}

        .v10-tooltip-grid {
            display:grid;
            grid-template-columns:repeat(4,minmax(0,1fr));
            gap:12px;
            margin:14px 0;
        }
        .v10-tooltip {
            background:#fffdf6;
            border:1px solid #f3d9a7;
            border-radius:18px;
            padding:14px;
        }
        .v10-tooltip-title {
            font-size:13px;
            font-weight:950;
            color:#7a4b00;
        }
        .v10-tooltip-text {
            font-size:12px;
            color:#62420d;
            line-height:1.45;
            margin-top:6px;
        }

        .v10-premium-divider {
            height:1px;
            background:linear-gradient(90deg,transparent,#9eb9dc,transparent);
            margin:20px 0;
        }

        @media (max-width:1150px) {
            .v10-action-grid {grid-template-columns:repeat(2,minmax(0,1fr));}
            .v10-quality-grid,.v10-control-room {grid-template-columns:1fr;}
            .v10-json-checks,.v10-tooltip-grid {grid-template-columns:repeat(2,minmax(0,1fr));}
        }
        @media (max-width:760px) {
            .v10-action-grid,.v10-mini-grid,.v10-json-checks,.v10-tooltip-grid {grid-template-columns:1fr;}
        }

    </style>
    """, unsafe_allow_html=True)



# =========================================================
# DATA PROCESSING
# =========================================================

def clean_header(value) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ").replace("\r", " ")
    text = text.replace("₹", "rs")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^a-z0-9%/() -]", "", text)
    return text.strip()


def find_col(df: pd.DataFrame, logical: str) -> Optional[str]:
    aliases = COLUMN_ALIASES.get(logical, [logical])
    norm = {clean_header(c): c for c in df.columns}
    for alias in aliases:
        ca = clean_header(alias)
        if ca in norm:
            return norm[ca]
    for alias in aliases:
        ca = clean_header(alias)
        for nc, orig in norm.items():
            if ca and ca in nc:
                return orig
    return None


def normalize_doc_no(x) -> str:
    text = str(x or "").strip().upper()
    text = re.sub(r"\.0$", "", text)
    return re.sub(r"[^A-Z0-9]", "", text)


def normalize_gstin(x) -> str:
    text = str(x or "").strip().upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def validate_gstin(x) -> bool:
    return bool(re.match(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]$", normalize_gstin(x)))


def to_number(s: pd.Series) -> pd.Series:
    if s is None:
        return pd.Series(dtype="float64")
    clean = (
        s.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("INR", "", case=False, regex=False)
        .str.replace("RS.", "", case=False, regex=False)
        .str.replace("RS", "", case=False, regex=False)
        .str.replace("₹", "", regex=False)
        .str.strip()
        .str.replace(r"^\((.*)\)$", r"-\1", regex=True)
    )
    clean = clean.replace({"": "0", "nan": "0", "None": "0", "NaT": "0"})
    return pd.to_numeric(clean, errors="coerce").fillna(0.0)


def to_date(s: pd.Series) -> pd.Series:
    if s is None:
        return pd.Series(dtype="datetime64[ns]")
    numeric = pd.to_numeric(s, errors="coerce")
    excel_dates = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
    parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)
    return parsed.fillna(excel_dates)


def sign_for_doc_type(x) -> int:
    text = str(x or "").lower()
    if any(k in text for k in ["credit", "cn", "cdn", "refund"]):
        return -1
    return 1


def detect_header_row(raw: pd.DataFrame) -> int:
    alias_words = [clean_header(a) for v in COLUMN_ALIASES.values() for a in v if len(clean_header(a)) >= 3]
    best_idx = raw.index[0]
    best_score = -1
    for idx, row in raw.head(30).iterrows():
        row_text = " ".join(clean_header(c) for c in row.tolist())
        score = sum(1 for a in alias_words if a and a in row_text)
        if score > best_score:
            best_score = score
            best_idx = idx
    return int(best_idx)


def normalize_sheet(sheet_df: pd.DataFrame) -> pd.DataFrame:
    raw = sheet_df.dropna(how="all").dropna(how="all", axis=1)
    if raw.empty:
        return pd.DataFrame()
    header_idx = detect_header_row(raw)
    header_pos = list(raw.index).index(header_idx)
    headers = []
    seen = {}
    for i, h in enumerate(raw.loc[header_idx].fillna("").astype(str).tolist(), start=1):
        name = str(h or "").strip() or f"Column {i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    body = raw.iloc[header_pos + 1:].copy()
    body.columns = headers
    body = body.dropna(how="all")
    body = body.loc[:, [c for c in body.columns if str(c).strip()]]
    return body


def read_excel_all_sheets(file, wanted_sheets: Optional[List[str]] = None) -> Dict[str, pd.DataFrame]:
    sheets = pd.read_excel(file, sheet_name=None, dtype=object, header=None, engine="openpyxl")
    out = {}
    for name, df in sheets.items():
        if wanted_sheets and name.strip().upper() not in [s.upper() for s in wanted_sheets]:
            continue
        clean = normalize_sheet(df)
        if not clean.empty:
            out[name.strip()] = clean
    return out


def standardize(df: pd.DataFrame, source_label: str, sheet_name: str = "", default_doc_type: str = "") -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()

    schema = {k: find_col(df, k) for k in COLUMN_ALIASES}
    n = len(df)

    def get(logical, default=""):
        col = schema.get(logical)
        if col and col in df.columns:
            return df[col]
        return pd.Series([default] * n, index=df.index)

    out = pd.DataFrame(index=df.index)
    out["supplier_gstin"] = get("supplier_gstin", "").map(normalize_gstin)
    out["supplier_name"] = get("supplier_name", "").astype(str).str.strip()
    out["document_type"] = get("document_type", default_doc_type).astype(str).replace({"": default_doc_type})
    out["document_no"] = get("document_no", "").astype(str).str.strip()
    out["document_norm"] = out["document_no"].map(normalize_doc_no)
    out["document_date"] = to_date(get("document_date", pd.NaT))
    out["invoice_value"] = to_number(get("invoice_value", 0))
    out["taxable_value"] = to_number(get("taxable_value", 0))
    out["igst"] = to_number(get("igst", 0))
    out["cgst"] = to_number(get("cgst", 0))
    out["sgst"] = to_number(get("sgst", 0))
    out["cess"] = to_number(get("cess", 0))
    out["total_tax"] = out[TAX_COLS].sum(axis=1)
    out["itc_available"] = get("itc_available", "Yes").astype(str)
    out["ims_status"] = get("ims_status", "No Action").astype(str)
    out["remarks"] = get("remarks", "").astype(str)
    out["pos"] = get("pos", "").astype(str)
    out["return_period"] = get("return_period", "").astype(str)
    out["source"] = source_label
    out["ims_sheet"] = sheet_name
    out["gstin_valid"] = out["supplier_gstin"].map(validate_gstin)
    out["data_quality"] = out.apply(row_quality_score, axis=1)

    sign = out["document_type"].map(sign_for_doc_type)
    for c in MONEY_COLS + ["total_tax"]:
        out[c] = out[c] * sign.where(out[c] >= 0, 1)

    out = out[
        ["source", "ims_sheet", "supplier_gstin", "supplier_name", "document_type", "document_no",
         "document_norm", "document_date", "invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess",
         "total_tax", "itc_available", "ims_status", "remarks", "pos", "return_period", "gstin_valid",
         "data_quality"]
    ]
    out = out[(out["supplier_gstin"].astype(str).str.len() > 0) | (out["document_norm"].astype(str).str.len() > 0)]
    return out.reset_index(drop=True)


def row_quality_score(row) -> int:
    score = 100
    if not validate_gstin(row.get("supplier_gstin", "")) and str(row.get("supplier_gstin", "")).strip():
        score -= 25
    if not str(row.get("document_norm", "")).strip():
        score -= 30
    if pd.isna(pd.to_datetime(row.get("document_date"), errors="coerce")):
        score -= 15
    if abs(float(row.get("taxable_value", 0) or 0)) < 0.01 and abs(float(row.get("total_tax", 0) or 0)) > 0.01:
        score -= 20
    return max(0, min(100, score))


def read_purchase_file(file) -> pd.DataFrame:
    if file is None:
        return pd.DataFrame()
    suffix = Path(file.name).suffix.lower()
    if suffix == ".csv":
        raw = pd.read_csv(file, dtype=object)
        return standardize(raw, "Purchase Register")
    sheets = read_excel_all_sheets(file)
    frames = []
    for sheet, raw in sheets.items():
        std = standardize(raw, "Purchase Register", sheet)
        if not std.empty:
            frames.append(std)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def infer_doc_type_from_sheet(sheet: str) -> str:
    s = sheet.upper()
    if "CN" in s:
        return "Credit Note"
    if "DN" in s:
        return "Debit Note"
    if "A" in s and "B2B" in s:
        return "Amended Invoice"
    return "Invoice"


def read_ims_utility(file) -> pd.DataFrame:
    if file is None:
        return pd.DataFrame()
    sheets = read_excel_all_sheets(file, IMS_SHEETS)
    frames = []
    for sheet, raw in sheets.items():
        default_doc = infer_doc_type_from_sheet(sheet)
        std = standardize(raw, "IMS Utility", sheet, default_doc)
        if not std.empty:
            frames.append(std)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def flatten_json(obj, parent_key="", sep="_"):
    items = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else str(k)
            items.extend(flatten_json(v, new_key, sep=sep).items())
    elif isinstance(obj, list):
        if all(isinstance(i, dict) for i in obj):
            return {parent_key: obj}
        return {parent_key: obj}
    else:
        return {parent_key: obj}
    return dict(items)


def extract_records_from_json(obj) -> List[dict]:
    records = []
    def walk(x, path=""):
        if isinstance(x, list):
            if x and all(isinstance(i, dict) for i in x):
                for item in x:
                    flat = flatten_json(item)
                    flat["_json_path"] = path
                    records.append(flat)
            else:
                for i, item in enumerate(x):
                    walk(item, f"{path}.{i}")
        elif isinstance(x, dict):
            for k, v in x.items():
                walk(v, f"{path}.{k}" if path else k)
    walk(obj)
    return records


def ims_json_records(data) -> List[dict]:
    section_map = {
        "b2b": "B2B", "b2ba": "B2BA", "b2bdn": "B2B-DN", "b2bdna": "B2B-DNA",
        "b2bcn": "B2B-CN", "b2bcna": "B2B-CNA", "cdnr": "B2B-CN", "cdnra": "B2B-CNA",
        "dn": "B2B-DN", "dna": "B2B-DNA", "cn": "B2B-CN", "cna": "B2B-CNA",
        "eco": "ECO", "ecoa": "ECOA"
    }
    found = []

    def walk(obj, path=""):
        if isinstance(obj, dict):
            for k, v in obj.items():
                key = str(k).lower().replace("_", "").replace("-", "")
                if key in section_map and isinstance(v, list):
                    for item in v:
                        if isinstance(item, dict):
                            rec = flatten_json(item)
                            rec["__section"] = section_map[key]
                            rec["__json_key"] = str(k)
                            found.append(rec)
                else:
                    walk(v, f"{path}.{k}" if path else str(k))
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                walk(item, f"{path}.{i}")

    walk(data)
    return found


def read_ims_json(file) -> pd.DataFrame:
    if file is None:
        return pd.DataFrame()
    data = json.load(file)
    records = ims_json_records(data)
    if not records:
        records = extract_records_from_json(data)
    if not records:
        return pd.DataFrame()
    raw = pd.DataFrame(records)
    section = raw.get("__section", pd.Series(["JSON"] * len(raw)))
    frames = []
    for sec, part in raw.groupby(section, dropna=False):
        default_doc = infer_doc_type_from_sheet(str(sec))
        std = standardize(part, "IMS JSON", str(sec), default_doc)
        if not std.empty:
            frames.append(std)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def parse_ims_json_bytes(file_bytes: bytes) -> Tuple[pd.DataFrame, List[dict], dict]:
    data = json.loads(file_bytes.decode("utf-8-sig"))
    records = ims_json_records(data)
    raw = pd.DataFrame(records) if records else pd.DataFrame()
    if raw.empty:
        return pd.DataFrame(), [], data
    frames = []
    for sec, part in raw.groupby(raw["__section"], dropna=False):
        std = standardize(part, "IMS JSON", str(sec), infer_doc_type_from_sheet(str(sec)))
        if not std.empty:
            frames.append(std)
    df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    return df, records, data


def to_excel_date_string(value):
    if value in [None, "", pd.NaT]:
        return ""
    try:
        dt = pd.to_datetime(value, dayfirst=True, errors="coerce")
        if pd.isna(dt):
            return str(value)
        return dt.strftime("%d-%m-%Y")
    except Exception:
        return str(value)


def get_json_value(rec: dict, *keys, default=""):
    for key in keys:
        if key in rec and rec.get(key) not in [None, ""]:
            return rec.get(key)
    lowered = {str(k).lower(): v for k, v in rec.items()}
    for key in keys:
        k = str(key).lower()
        if k in lowered and lowered[k] not in [None, ""]:
            return lowered[k]
    return default


def action_for_record(rec: dict, action_map: dict) -> str:
    gstin = normalize_gstin(get_json_value(rec, "stin", "ctin", "supplier_gstin"))
    doc_no = normalize_doc_no(get_json_value(rec, "inum", "nt_num", "document_no", "oinum"))
    return action_map.get((gstin, doc_no), "Pending")


def build_action_map(recon: pd.DataFrame) -> dict:
    if recon is None or recon.empty:
        return {}
    out = {}
    for _, row in recon.iterrows():
        gstin = normalize_gstin(row.get("supplier_gstin", ""))
        doc = normalize_doc_no(row.get("document_norm", ""))
        action = "Accepted" if str(row.get("mismatch_type", "")) == "Matched" else "Pending"
        if gstin and doc:
            out[(gstin, doc)] = action
    return out


def clear_utility_rows(ws, start_row: int, max_col: int):
    last = max(ws.max_row, start_row + 500)
    for row in ws.iter_rows(min_row=start_row, max_row=last, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None


def write_row_values(ws, row: int, values: dict):
    for col, value in values.items():
        ws.cell(row=row, column=col).value = value


def populate_ims_utility_xlsm(template_bytes: bytes, records: List[dict], recon: pd.DataFrame) -> bytes:
    if not template_bytes:
        raise ValueError("Please upload Inbuilt IMS Utility .xlsm template first.")
    if not records:
        raise ValueError("Please upload/process IMS JSON first.")

    wb = load_workbook(BytesIO(template_bytes), keep_vba=True)
    action_map = build_action_map(recon)
    groups: Dict[str, List[dict]] = {}
    for rec in records:
        groups.setdefault(str(rec.get("__section", "B2B")), []).append(rec)

    config = {
        "B2B": {"start": 7, "max_col": 23, "amend": False, "note": False},
        "B2B-DN": {"start": 7, "max_col": 23, "amend": False, "note": False},
        "B2B-CN": {"start": 7, "max_col": 35, "amend": False, "note": True},
        "B2BA": {"start": 8, "max_col": 37, "amend": True, "note": False},
        "B2B-DNA": {"start": 8, "max_col": 37, "amend": True, "note": False},
        "B2B-CNA": {"start": 8, "max_col": 37, "amend": True, "note": True},
    }

    for sheet_name, rows in groups.items():
        if sheet_name not in wb.sheetnames or sheet_name not in config:
            continue
        ws = wb[sheet_name]
        cfg = config[sheet_name]
        start = cfg["start"]
        clear_utility_rows(ws, start, cfg["max_col"])

        for idx, rec in enumerate(rows, start=start):
            status = action_for_record(rec, action_map)
            gstin = normalize_gstin(get_json_value(rec, "stin", "ctin", "supplier_gstin"))
            trade = get_json_value(rec, "tradenm", "supplier_name")
            doc_no = get_json_value(rec, "inum", "nt_num", "document_no")
            doc_type = get_json_value(rec, "inv_typ", "ntty", "document_type", default="R")
            doc_date = to_excel_date_string(get_json_value(rec, "idt", "nt_dt", "document_date"))
            doc_val = get_json_value(rec, "val", "invoice_value", default=0)
            pos = get_json_value(rec, "pos", default="")
            txval = get_json_value(rec, "txval", "taxable_value", default=0)
            iamt = get_json_value(rec, "iamt", "igst", default=0)
            camt = get_json_value(rec, "camt", "cgst", default=0)
            samt = get_json_value(rec, "samt", "sgst", default=0)
            cess = get_json_value(rec, "cess", default=0)
            remarks = "Auto Accepted by IMS Recon Pro" if status == "Accepted" else "Auto Pending - not matched in Purchase Register"
            src = get_json_value(rec, "srcform", default="")
            rtnprd = get_json_value(rec, "rtnprd", default="")
            filing = get_json_value(rec, "srcfilstatus", default="")
            pending_block = get_json_value(rec, "ispendactblocked", default="N")
            remarks_block = get_json_value(rec, "isRemarksBlocked", "isremarksblocked", default="N")

            if not cfg["amend"]:
                values = {
                    1: gstin, 2: trade, 3: doc_no, 4: doc_type, 5: doc_date, 6: float(doc_val or 0),
                    7: status, 8: pos, 9: float(txval or 0), 10: float(iamt or 0), 11: float(camt or 0),
                    12: float(samt or 0), 13: float(cess or 0),
                    14: remarks if sheet_name != "B2B-CN" else "No", 15: src, 16: rtnprd, 17: filing,
                    20: get_json_value(rec, "action", default="N"),
                }
                if sheet_name == "B2B-CN":
                    values.update({19: remarks, 20: src, 21: rtnprd, 22: filing, 25: get_json_value(rec, "action", default="N"), 32: pending_block, 33: remarks_block})
                else:
                    values.update({22: pending_block, 23: remarks_block})
            else:
                orig_no = get_json_value(rec, "oinum", "org_inum", "oinv_num", default="")
                orig_dt = to_excel_date_string(get_json_value(rec, "oidt", "org_idt", "oinv_dt", default=""))
                values = {
                    1: orig_no, 2: orig_dt, 3: gstin, 4: trade, 5: doc_no, 6: doc_type, 7: doc_date, 8: float(doc_val or 0),
                    9: status, 10: pos, 11: float(txval or 0), 12: float(iamt or 0), 13: float(camt or 0),
                    14: float(samt or 0), 15: float(cess or 0), 16: "No", 21: remarks, 22: src,
                    23: rtnprd, 24: filing, 27: get_json_value(rec, "action", default="N"),
                    34: pending_block, 35: remarks_block
                }
            write_row_values(ws, idx, values)
            fill = PatternFill("solid", fgColor="E2F0D9") if status == "Accepted" else PatternFill("solid", fgColor="FFF2CC")
            ws.cell(idx, 7 if not cfg["amend"] else 9).fill = fill
            ws.cell(idx, 7 if not cfg["amend"] else 9).font = Font(bold=True)

    if "Home" in wb.sheetnames:
        ws = wb["Home"]
        try:
            rtin = ""
            for rec in records:
                rtin = get_json_value(rec, "rtin", default="") or rtin
            ws["B5"] = "GSTIN"
            ws["C5"] = rtin or st.session_state.get("client_gstin", "")
        except Exception:
            pass

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def normalize_action_label(value) -> str:
    """Convert IMS Utility status/dropdown value into one of Accepted/Pending/Rejected/No Action."""
    text = str(value or "").strip().lower()
    if text in ["a", "accept", "accepted"]:
        return "Accepted"
    if text in ["p", "pending"]:
        return "Pending"
    if text in ["r", "reject", "rejected"]:
        return "Rejected"
    if text in ["n", "no", "no action", "noaction", "na", ""]:
        return "No Action"
    if "accept" in text:
        return "Accepted"
    if "pend" in text or "review" in text:
        return "Pending"
    if "reject" in text:
        return "Rejected"
    return "Pending"


def action_label_to_gst_code(action: str) -> str:
    """GST IMS JSON action code.

    Important for GST upload JSON:
    - Accepted -> A
    - Pending  -> P
    - Rejected -> R
    - No Action -> N internally only.

    In the final GST upload file, records with No Action/N are skipped, because
    the official utility output sample contains actioned records only.
    """
    label = normalize_action_label(action)
    return {"Accepted": "A", "Pending": "P", "Rejected": "R", "No Action": "N"}.get(label, "P")


def utility_sheet_config() -> Dict[str, dict]:
    return {
        "B2B": {"start": 7, "gstin_col": 1, "doc_col": 3, "status_col": 7},
        "B2B-DN": {"start": 7, "gstin_col": 1, "doc_col": 3, "status_col": 7},
        "B2B-CN": {"start": 7, "gstin_col": 1, "doc_col": 3, "status_col": 7},
        "B2BA": {"start": 8, "gstin_col": 3, "doc_col": 5, "status_col": 9},
        "B2B-DNA": {"start": 8, "gstin_col": 3, "doc_col": 5, "status_col": 9},
        "B2B-CNA": {"start": 8, "gstin_col": 3, "doc_col": 5, "status_col": 9},
    }


def read_action_status_from_utility_xlsm(xlsm_bytes: bytes) -> Tuple[dict, pd.DataFrame]:
    """Read final action/status selected by user from Inbuilt IMS Utility .xlsm."""
    if not xlsm_bytes:
        raise ValueError("Please upload the final/edited IMS JSON file first.")
    wb = load_workbook(BytesIO(xlsm_bytes), keep_vba=True, data_only=False)
    action_map = {}
    rows = []

    for sheet_name, cfg in utility_sheet_config().items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        blank_streak = 0
        for r in range(cfg["start"], ws.max_row + 1):
            gstin = normalize_gstin(ws.cell(r, cfg["gstin_col"]).value)
            doc_no_raw = ws.cell(r, cfg["doc_col"]).value
            doc_norm = normalize_doc_no(doc_no_raw)
            status_raw = ws.cell(r, cfg["status_col"]).value
            status = normalize_action_label(status_raw)

            if not gstin and not doc_norm:
                blank_streak += 1
                if blank_streak >= 25:
                    break
                continue
            blank_streak = 0
            if not doc_norm:
                continue

            action_map[(sheet_name, gstin, doc_norm)] = status
            action_map[(gstin, doc_norm)] = status
            rows.append({
                "Sheet": sheet_name,
                "Supplier GSTIN": gstin,
                "Document No": doc_no_raw,
                "Normalized Document No": doc_norm,
                "Utility Status": status,
                "GST JSON Action Code": action_label_to_gst_code(status),
            })

    summary = pd.DataFrame(rows)
    return action_map, summary


def get_json_section_map() -> dict:
    """
    Map downloaded IMS section names to the exact upload JSON section keys.

    Important: GST portal upload schema is case-sensitive. The official utility output
    uses lowercase keys inside `invdata` like `b2b`, not display names like `B2B`.
    """
    return {
        "b2b": "b2b",
        "b2ba": "b2ba",
        "b2bdn": "b2bdn",
        "b2bdna": "b2bdna",
        "b2bcn": "b2bcn",
        "b2bcna": "b2bcna",
        "cdnr": "b2bcn",
        "cdnra": "b2bcna",
        "dn": "b2bdn",
        "dna": "b2bdna",
        "cn": "b2bcn",
        "cna": "b2bcna",
        "eco": "eco",
        "ecoa": "ecoa",
    }


def update_ims_json_actions_from_utility(original_json: dict, action_map: dict) -> Tuple[dict, pd.DataFrame]:
    """Preserve GST portal JSON structure and update only action field based on utility status."""
    if not isinstance(original_json, dict) or not original_json:
        raise ValueError("Original IMS JSON is not available. Please process IMS JSON first.")
    if not action_map:
        raise ValueError("No status/action found in the uploaded IMS JSON.")

    data = deepcopy(original_json)
    section_map = get_json_section_map()
    updated_rows = []

    def walk(obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                normalized_key = str(k).lower().replace("_", "").replace("-", "")
                if normalized_key in section_map and isinstance(v, list):
                    section = section_map[normalized_key]
                    for item in v:
                        if not isinstance(item, dict):
                            continue
                        gstin = normalize_gstin(get_json_value(item, "stin", "ctin", "supplier_gstin"))
                        doc_no = get_json_value(item, "inum", "nt_num", "document_no", "oinum")
                        doc_norm = normalize_doc_no(doc_no)
                        status = action_map.get((section, gstin, doc_norm)) or action_map.get((gstin, doc_norm))
                        if status:
                            item["action"] = action_label_to_gst_code(status)
                            if "remarks" in item and not str(item.get("remarks") or "").strip():
                                item["remarks"] = f"Auto {status} by IMS Recon Pro"
                            updated_rows.append({
                                "Sheet": section,
                                "Supplier GSTIN": gstin,
                                "Document No": doc_no,
                                "Utility Status": status,
                                "GST JSON Action Code": item.get("action", ""),
                            })
                else:
                    walk(v)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    walk(data)
    return data, pd.DataFrame(updated_rows)


def generate_gst_upload_json_bytes(original_json: dict, final_xlsm_bytes: bytes) -> Tuple[bytes, pd.DataFrame, pd.DataFrame]:
    action_map, utility_status_df = read_action_status_from_utility_xlsm(final_xlsm_bytes)
    updated_json, updated_summary = update_ims_json_actions_from_utility(original_json, action_map)
    json_bytes = json.dumps(updated_json, ensure_ascii=False, indent=2).encode("utf-8")
    return json_bytes, utility_status_df, updated_summary



def _compact_json_number(value):
    """Return Python number/string without converting zero/decimal fields into invalid schema values."""
    if value is None:
        return value
    # bool is subclass of int; do not treat it as number here.
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        try:
            f = float(value)
            if f.is_integer():
                return int(f)
            return round(f, 2)
        except Exception:
            return value
    return value


def _clean_ims_upload_record(source_item: dict, action_code: str) -> dict:
    """
    Create one GST IMS upload record.

    V6 important correction:
    - For normal B2B records, the official utility output has common keys.
    - For amendment sections such as b2ba/b2bdna/b2bcna, GST may require
      additional original-document fields from the downloaded JSON.
    - Therefore we preserve all GST-source keys from the original downloaded
      JSON record except known non-upload/display/internal keys.
    - This prevents amendment records from losing mandatory fields and causing
      GST schema validation failure.
    """

    if not isinstance(source_item, dict):
        source_item = {}

    # Never send these to GST upload JSON. They are either display-only from the
    # downloaded JSON or internal to this app/reporting layer.
    blocked_keys = {
        "tradenm", "tradeNm", "trade_name", "supplier_name", "hash",
        "remarks", "remark", "comments", "comment",
        "recommended_action", "final_user_action", "user_remarks",
        "mismatch_type", "risk_level", "risk_score", "reason",
        "match_status", "confidence_score", "source", "ims_sheet",
        "document_norm", "data_quality", "gstin_valid",
    }

    # Keep GST utility's familiar order first, then append amendment-specific
    # keys from the source record, preserving their original names.
    preferred_order = [
        "stin", "inum", "inv_typ", "idt", "val", "action", "pos", "txval",
        "iamt", "camt", "samt", "cess", "srcform", "rtnprd", "srcfilstatus",
        "ispendactblocked", "isRemarksBlocked"
    ]

    out = {}
    for key in preferred_order:
        if key == "action":
            out["action"] = action_code
        elif key in source_item and key not in blocked_keys:
            out[key] = _compact_json_number(source_item.get(key))

    # Append remaining GST-source keys, e.g. original invoice/date fields in
    # amendment sections. Do not invent names; keep exactly what GST JSON gave us.
    for key, value in source_item.items():
        if key in blocked_keys or key in out or key == "action":
            continue
        # Avoid app-created Python helper columns accidentally entering JSON.
        if str(key).startswith("_"):
            continue
        out[key] = _compact_json_number(value)

    # Apply safe defaults only for common numeric/flag fields.
    for tax_key in ["iamt", "camt", "samt", "cess"]:
        out.setdefault(tax_key, 0)
    out.setdefault("ispendactblocked", "N")
    out.setdefault("isRemarksBlocked", "N")
    out["action"] = action_code
    return out


def _record_missing_mandatory(upload_item: dict) -> list:
    mandatory = ["stin", "inum", "inv_typ", "idt", "val", "action", "pos", "txval", "srcform", "rtnprd", "srcfilstatus"]
    missing = []
    for key in mandatory:
        if key not in upload_item or upload_item.get(key) in [None, ""]:
            missing.append(key)
    return missing



def ims_json_section_counts(original_json: dict) -> pd.DataFrame:
    """Return section-wise counts from the uploaded GST IMS JSON / generated upload JSON."""
    section_map = get_json_section_map()
    rows = []
    if not isinstance(original_json, dict) or not original_json:
        return pd.DataFrame(columns=["Section", "Records"])
    root = original_json.get("imsDetails", original_json.get("invdata", original_json))

    def normalized_section_name(key: str) -> str:
        raw = str(key or "").strip()
        norm = raw.lower().replace("_", "").replace("-", "")
        return section_map.get(norm, raw.lower())

    def walk(obj):
        if isinstance(obj, dict):
            for key, value in obj.items():
                norm = str(key).lower().replace("_", "").replace("-", "")
                if norm in section_map and isinstance(value, list):
                    rows.append({"Section": normalized_section_name(key), "Records": len(value)})
                elif isinstance(value, (dict, list)):
                    walk(value)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    walk(root)
    if not rows:
        return pd.DataFrame(columns=["Section", "Records"])
    out = pd.DataFrame(rows).groupby("Section", as_index=False)["Records"].sum()
    return out.sort_values("Section").reset_index(drop=True)


def generated_json_action_counts(json_bytes: bytes) -> pd.DataFrame:
    """Return section/action counts from the final GST upload JSON bytes."""
    try:
        data = json.loads(json_bytes.decode("utf-8") if isinstance(json_bytes, (bytes, bytearray)) else str(json_bytes))
    except Exception:
        return pd.DataFrame(columns=["Section", "Action", "Records"])
    rows = []
    invdata = data.get("invdata", {}) if isinstance(data, dict) else {}
    if isinstance(invdata, dict):
        for section, records in invdata.items():
            if isinstance(records, list):
                for rec in records:
                    if isinstance(rec, dict):
                        rows.append({"Section": section, "Action": rec.get("action", ""), "Records": 1})
    if not rows:
        return pd.DataFrame(columns=["Section", "Action", "Records"])
    return pd.DataFrame(rows).groupby(["Section", "Action"], as_index=False)["Records"].sum()

def generate_gst_upload_json_from_final_actions(original_json: dict, action_df: pd.DataFrame) -> Tuple[bytes, pd.DataFrame]:
    """
    Generate GST Portal IMS upload JSON in official utility style.

    The user's official output sample confirms this wrapper:
    {
      "rtin": "...",
      "reqtyp": "SAVE",
      "invdata": {"b2b": [ ... ]}
    }

    This V6 generator is amendment-safe and GST-utility style:
    - lowercase invdata section keys
    - no imsDetails wrapper
    - no tradenm/hash/remarks/internal fields, but preserves amendment mandatory fields
    - every original IMS JSON record is actively actioned by default
    - Matched/manual Accepted => A
    - Unmatched/missing/No Action/Review => P (Pending), so records do not remain No Action on portal
    - invalid identity records are skipped and shown in summary
    """
    if not isinstance(original_json, dict) or not original_json:
        raise ValueError("Original IMS JSON is not available. Please process IMS JSON first.")
    if action_df is None or action_df.empty:
        raise ValueError("Final action table is empty. Please run reconciliation and save final actions first.")

    action_map = {}
    for _, row in action_df.iterrows():
        gstin = normalize_gstin(row.get("supplier_gstin", ""))
        doc_norm = normalize_doc_no(row.get("document_norm", ""))
        if not gstin or not doc_norm:
            continue
        action = normalize_action_label(row.get("final_user_action", row.get("recommended_action", "Pending")))
        section = str(row.get("ims_sheet_ims", row.get("ims_sheet", "")) or "").strip().lower()
        action_map[(gstin, doc_norm)] = action
        if section:
            action_map[(section, gstin, doc_norm)] = action

    section_map = get_json_section_map()
    invdata = {}
    updated_rows = []
    skipped_rows = []

    source_root = original_json.get("imsDetails", original_json.get("invdata", original_json))

    def normalized_section_name(key: str) -> str:
        raw = str(key or "").strip()
        norm = raw.lower().replace("_", "").replace("-", "")
        return section_map.get(norm, raw.lower())

    def process_rows(section_key: str, rows: list):
        section = normalized_section_name(section_key)
        if not isinstance(rows, list):
            return

        for item in rows:
            if not isinstance(item, dict):
                continue

            gstin = normalize_gstin(get_json_value(item, "stin", "ctin", "supplier_gstin"))
            doc_no = get_json_value(item, "inum", "nt_num", "document_no", "oinum")
            doc_norm = normalize_doc_no(doc_no)
            if not gstin or not doc_norm:
                skipped_rows.append({"Section": section, "Supplier GSTIN": gstin, "Document No": doc_no, "Reason": "Missing GSTIN or document number"})
                continue

            status = (
                action_map.get((section, gstin, doc_norm))
                or action_map.get((str(section_key).lower(), gstin, doc_norm))
                or action_map.get((gstin, doc_norm))
            )

            # Very important GST IMS rule for this project:
            # after reconciliation, every IMS invoice should be actioned.
            # Matched records become Accepted. All records not found/mapped in the
            # Action Center are treated as Pending, not No Action, because the
            # portal otherwise continues to show them under No Action.
            if not status:
                status = "Pending"

            action_code = action_label_to_gst_code(status)

            # For GST upload, do not leave JSON records as N/No Action.
            # If the app/user action is No Action or Review, keep it Pending so
            # the portal moves it out of No Action and the user can take final
            # action later from IMS dashboard if required.
            if action_code == "N":
                status = "Pending"
                action_code = "P"

            upload_item = _clean_ims_upload_record(item, action_code)
            missing = _record_missing_mandatory(upload_item)
            if missing:
                skipped_rows.append({"Section": section, "Supplier GSTIN": gstin, "Document No": doc_no, "Reason": "Missing mandatory fields: " + ", ".join(missing)})
                continue

            invdata.setdefault(section, []).append(upload_item)
            updated_rows.append({
                "Section": section,
                "Supplier GSTIN": gstin,
                "Document No": doc_no,
                "Final Action": status,
                "GST JSON Action Code": action_code,
                "Validation": "Included in GST upload JSON",
            })

    def walk_sections(obj):
        if isinstance(obj, dict):
            for key, value in obj.items():
                norm = str(key).lower().replace("_", "").replace("-", "")
                if norm in section_map and isinstance(value, list):
                    process_rows(key, value)
                elif isinstance(value, (dict, list)):
                    walk_sections(value)
        elif isinstance(obj, list):
            for item in obj:
                walk_sections(item)

    walk_sections(source_root)

    # Remove any empty section defensively.
    invdata = {k: v for k, v in invdata.items() if isinstance(v, list) and len(v) > 0}

    if not invdata:
        raise ValueError("No valid IMS records found for GST JSON generation. Please check the uploaded GST IMS JSON.")

    upload_json = {
        "rtin": str(original_json.get("rtin", st.session_state.get("client_gstin", ""))).strip(),
        "reqtyp": "SAVE",
        "invdata": invdata,
    }

    json_bytes = json.dumps(upload_json, ensure_ascii=False, indent=3).encode("utf-8")
    included = pd.DataFrame(updated_rows)
    skipped = pd.DataFrame(skipped_rows)
    if not skipped.empty:
        skipped.insert(0, "GST JSON Action Code", "SKIPPED")
    summary = pd.concat([included, skipped], ignore_index=True) if not skipped.empty else included
    return json_bytes, summary



# =========================================================
# V7 STRONGER RECONCILIATION + VALIDATION HELPERS
# =========================================================

def safe_float_value(value) -> float:
    try:
        if pd.isna(value):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def approx_equal(a, b, tolerance: float) -> bool:
    return abs(safe_float_value(a) - safe_float_value(b)) <= float(tolerance or 0)


def date_gap_days(a, b) -> Optional[int]:
    da = pd.to_datetime(a, errors="coerce")
    db = pd.to_datetime(b, errors="coerce")
    if pd.isna(da) or pd.isna(db):
        return None
    return int(abs((da - db).days))


def is_credit_note_text(value) -> bool:
    text = str(value or "").lower()
    return any(x in text for x in ["credit", "cn", "cdn", "b2b-cn", "b2bcna"])


def make_recon_key(gstin, doc_norm) -> str:
    return f"{normalize_gstin(gstin)}|{normalize_doc_no(doc_norm)}"


def enhance_recon_row(row, amount_tol: float, date_tol: int) -> pd.Series:
    """Classify exact key matches using amount, tax-head and date checks."""
    merge_status = str(row.get("_merge", ""))
    if merge_status == "left_only":
        row["mismatch_type"] = "Only in Purchase Register"
        row["match_level"] = "L0 Not in IMS"
        return row
    if merge_status == "right_only":
        row["mismatch_type"] = "Only in IMS"
        row["match_level"] = "L0 Not in Purchase Register"
        return row

    amount_ok = (
        abs(safe_float_value(row.get("taxable_value_diff"))) <= amount_tol
        and abs(safe_float_value(row.get("total_tax_diff"))) <= amount_tol
        and abs(safe_float_value(row.get("invoice_value_diff"))) <= max(amount_tol, 1)
    )
    tax_head_ok = all(abs(safe_float_value(row.get(f"{c}_diff"))) <= amount_tol for c in ["igst", "cgst", "sgst", "cess"])
    gap = row.get("date_diff_days")
    try:
        date_ok = int(gap) <= int(date_tol)
    except Exception:
        date_ok = True

    row["match_level"] = "L1 Exact GSTIN + Invoice No"
    if amount_ok and tax_head_ok and date_ok:
        row["mismatch_type"] = "Matched"
    elif amount_ok and not tax_head_ok:
        row["mismatch_type"] = "Tax Head Mismatch"
    elif (not amount_ok) and date_ok:
        row["mismatch_type"] = "Value / Tax Mismatch"
    elif amount_ok and tax_head_ok and not date_ok:
        row["mismatch_type"] = "Date Mismatch"
    else:
        row["mismatch_type"] = "Value and Date Mismatch"
    return row


def add_probable_match_flags(result: pd.DataFrame, p_agg: pd.DataFrame, i_agg: pd.DataFrame, amount_tol: float, date_tol: int) -> pd.DataFrame:
    """Mark IMS-only rows where supplier/value/date indicates a likely invoice-number difference."""
    if result.empty or p_agg.empty or i_agg.empty:
        return result
    p_lookup = p_agg.copy()
    i_only_mask = result["_merge"].astype(str).eq("right_only")
    if not i_only_mask.any():
        return result

    for idx, row in result[i_only_mask].iterrows():
        gstin = row.get("supplier_gstin", "")
        candidates = p_lookup[p_lookup["supplier_gstin"].astype(str).eq(str(gstin))].copy()
        if candidates.empty:
            continue
        best_score = 999999.0
        best = None
        for _, p in candidates.iterrows():
            tax_gap = abs(safe_float_value(p.get("total_tax_purchase")) - safe_float_value(row.get("total_tax_ims")))
            taxable_gap = abs(safe_float_value(p.get("taxable_value_purchase")) - safe_float_value(row.get("taxable_value_ims")))
            inv_gap = abs(safe_float_value(p.get("invoice_value_purchase")) - safe_float_value(row.get("invoice_value_ims")))
            dg = date_gap_days(p.get("document_date_purchase"), row.get("document_date_ims"))
            dg_score = 9999 if dg is None else dg
            if taxable_gap <= amount_tol and tax_gap <= amount_tol and inv_gap <= max(amount_tol, 1) and dg_score <= date_tol:
                score = taxable_gap + tax_gap + inv_gap + dg_score
                if score < best_score:
                    best_score = score
                    best = p
        if best is not None:
            result.loc[idx, "mismatch_type"] = "Probable Match - Invoice No Difference"
            result.loc[idx, "match_level"] = "L3 GSTIN + Date + Value"
            result.loc[idx, "confidence_score"] = 78
            result.loc[idx, "reason"] = "GSTIN, date and values are close but invoice/document number differs. Review invoice number format before accepting."
            for col in ["supplier_name_purchase", "document_type_purchase", "document_no_purchase", "document_date_purchase", "invoice_value_purchase", "taxable_value_purchase", "igst_purchase", "cgst_purchase", "sgst_purchase", "cess_purchase", "total_tax_purchase"]:
                if col in best.index:
                    result.loc[idx, col] = best[col]
            for c in MONEY_COLS + ["total_tax"]:
                result.loc[idx, f"{c}_diff"] = safe_float_value(result.loc[idx].get(f"{c}_purchase")) - safe_float_value(result.loc[idx].get(f"{c}_ims"))
            result.loc[idx, "date_diff_days"] = date_gap_days(result.loc[idx].get("document_date_purchase"), result.loc[idx].get("document_date_ims")) or 0
    return result


def upload_quality_summary(df: pd.DataFrame, label: str) -> pd.DataFrame:
    """Upload validation with value columns for taxation review."""
    value_cols = ["taxable_value", "igst", "cgst", "sgst", "cess", "total_tax"]

    def blank_series(index=None):
        return pd.Series([False] * (0 if index is None else len(index)), index=index)

    def amount_row(check: str, subset: pd.DataFrame) -> dict:
        subset = subset if subset is not None else pd.DataFrame()
        row = {
            "Check": check,
            "Records": int(len(subset)),
            "Taxable Value": round(safe_float_value(subset.get("taxable_value", pd.Series(dtype=float)).sum()) if "taxable_value" in subset else 0, 2),
            "IGST": round(safe_float_value(subset.get("igst", pd.Series(dtype=float)).sum()) if "igst" in subset else 0, 2),
            "CGST": round(safe_float_value(subset.get("cgst", pd.Series(dtype=float)).sum()) if "cgst" in subset else 0, 2),
            "SGST": round(safe_float_value(subset.get("sgst", pd.Series(dtype=float)).sum()) if "sgst" in subset else 0, 2),
            "CESS": round(safe_float_value(subset.get("cess", pd.Series(dtype=float)).sum()) if "cess" in subset else 0, 2),
            "Total Tax": round(safe_float_value(subset.get("total_tax", pd.Series(dtype=float)).sum()) if "total_tax" in subset else 0, 2),
        }
        return row

    if df is None or df.empty:
        return pd.DataFrame([amount_row(f"{label} records", pd.DataFrame())])

    work = df.copy()
    rows = []
    rows.append(amount_row(f"{label} records", work))

    if "gstin_valid" in work.columns:
        valid_mask = work["gstin_valid"].fillna(False).astype(bool)
        rows.append(amount_row("Valid GSTIN", work[valid_mask]))
        rows.append(amount_row("Invalid GSTIN", work[~valid_mask]))
    else:
        rows.append(amount_row("Valid GSTIN", pd.DataFrame()))
        rows.append(amount_row("Invalid GSTIN", pd.DataFrame()))

    if "document_norm" in work.columns:
        blank_inv = work["document_norm"].astype(str).eq("")
        rows.append(amount_row("Blank invoice/document no", work[blank_inv]))
    else:
        rows.append(amount_row("Blank invoice/document no", pd.DataFrame()))

    if {"supplier_gstin", "document_norm"}.issubset(work.columns):
        dup_mask = work.duplicated(["supplier_gstin", "document_norm"], keep=False)
        rows.append(amount_row("Duplicate GSTIN + invoice/document no", work[dup_mask]))
    else:
        rows.append(amount_row("Duplicate GSTIN + invoice/document no", pd.DataFrame()))

    if "document_date" in work.columns:
        blank_date = pd.to_datetime(work["document_date"], errors="coerce").isna()
        rows.append(amount_row("Blank document date", work[blank_date]))
    else:
        rows.append(amount_row("Blank document date", pd.DataFrame()))

    # Additional tax-wise check helpful for GST reconciliation review
    if all(c in work.columns for c in ["igst", "cgst", "sgst"]):
        tax_blank = work[["igst", "cgst", "sgst"]].fillna(0).abs().sum(axis=1).le(0.009)
        rows.append(amount_row("Zero IGST/CGST/SGST", work[tax_blank]))

    return pd.DataFrame(rows)


def duplicate_report(df: pd.DataFrame, label: str) -> pd.DataFrame:
    if df is None or df.empty or not {"supplier_gstin", "document_norm"}.issubset(df.columns):
        return pd.DataFrame()
    dup = df[df.duplicated(["supplier_gstin", "document_norm"], keep=False)].copy()
    if dup.empty:
        return pd.DataFrame()
    cols = [
        c for c in [
            "supplier_gstin", "supplier_name", "document_no", "document_date",
            "invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess", "total_tax",
            "source", "ims_sheet"
        ] if c in dup.columns
    ]
    out = dup[cols].copy()
    rename_map = {
        "invoice_value": "Invoice Value",
        "taxable_value": "Taxable Value",
        "igst": "IGST",
        "cgst": "CGST",
        "sgst": "SGST",
        "cess": "CESS",
        "total_tax": "Total Tax",
        "supplier_gstin": "Supplier GSTIN",
        "supplier_name": "Supplier Name",
        "document_no": "Invoice/Document No",
        "document_date": "Document Date",
        "source": "Source",
        "ims_sheet": "IMS Section",
    }
    out = out.rename(columns=rename_map)
    out.insert(0, "Dataset", label)
    return out


def final_json_review_table(action_df: pd.DataFrame) -> pd.DataFrame:
    if action_df is None or action_df.empty:
        return pd.DataFrame()
    work = action_df.copy()
    work["GST JSON Code"] = work.get("final_user_action", "Pending").apply(action_label_to_gst_code)
    return work.groupby(["final_user_action", "GST JSON Code"], dropna=False).size().reset_index(name="Records")


def split_report_sheets(p: pd.DataFrame, ims: pd.DataFrame, recon: pd.DataFrame, action: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    sheets = {
        "Summary": recon_summary(recon),
        "Final Action Report": action,
        "Reconciliation": recon,
        "Purchase Standardized": p,
        "IMS JSON Standardized": ims,
        "Purchase Quality": upload_quality_summary(p, "Purchase Register"),
        "IMS Quality": upload_quality_summary(ims, "IMS JSON"),
        "Purchase Duplicates": duplicate_report(p, "Purchase Register"),
        "IMS Duplicates": duplicate_report(ims, "IMS JSON"),
        "Audit Log": load_audit(st.session_state.username),
    }
    if recon is not None and not recon.empty:
        sheets.update({
            "Matched": recon[recon["mismatch_type"].eq("Matched")],
            "Pending Cases": action[action.get("final_user_action", pd.Series()).eq("Pending")] if action is not None and not action.empty else pd.DataFrame(),
            "Rejected Cases": action[action.get("final_user_action", pd.Series()).eq("Rejected")] if action is not None and not action.empty else pd.DataFrame(),
            "Only in IMS": recon[recon["mismatch_type"].eq("Only in IMS")],
            "Only in Purchase": recon[recon["mismatch_type"].eq("Only in Purchase Register")],
            "Value Mismatch": recon[recon["mismatch_type"].astype(str).str.contains("Value|Tax Head", case=False, na=False)],
            "High Risk": recon[recon["risk_level"].isin(["High", "Critical"])],
            "Probable Matches": recon[recon["mismatch_type"].astype(str).str.contains("Probable", case=False, na=False)],
        })
    return sheets

def aggregate(df: pd.DataFrame, label: str) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    work = df.copy()
    work = work[(work["supplier_gstin"].astype(str) != "") & (work["document_norm"].astype(str) != "")]
    if work.empty:
        return pd.DataFrame()

    agg_map = {
        "supplier_name": "first",
        "document_type": "first",
        "document_no": "first",
        "document_date": "min",
        "invoice_value": "sum",
        "taxable_value": "sum",
        "igst": "sum",
        "cgst": "sum",
        "sgst": "sum",
        "cess": "sum",
        "total_tax": "sum",
        "itc_available": "first",
        "ims_status": "first",
        "remarks": "first",
        "ims_sheet": "first",
        "data_quality": "mean",
    }
    out = work.groupby(["supplier_gstin", "document_norm"], dropna=False).agg(agg_map).reset_index()
    rename = {col: f"{col}_{label}" for col in out.columns if col not in ["supplier_gstin", "document_norm"]}
    return out.rename(columns=rename)


def calculate_recon(purchase: pd.DataFrame, ims: pd.DataFrame, amount_tol: float, date_tol: int, include_amendments: bool) -> pd.DataFrame:
    """V7 stronger reconciliation engine.

    Matching levels:
    L1: GSTIN + cleaned invoice/document number
    L3: GSTIN + date + values where invoice number differs (probable match flag)
    All exact matches are then tested for value/date/tax-head mismatch.
    """
    if purchase is None or ims is None or purchase.empty or ims.empty:
        return pd.DataFrame()

    ims_work = ims.copy()
    if not include_amendments and "ims_sheet" in ims_work.columns:
        ims_work = ims_work[~ims_work["ims_sheet"].astype(str).str.upper().isin(["B2BA", "B2B-DNA", "B2B-CNA", "B2BDN", "B2BCN"])]

    p = aggregate(purchase, "purchase")
    i = aggregate(ims_work, "ims")
    if p.empty or i.empty:
        return pd.DataFrame()

    m = p.merge(i, on=["supplier_gstin", "document_norm"], how="outer", indicator=True)

    for c in MONEY_COLS + ["total_tax"]:
        m[f"{c}_diff"] = m.get(f"{c}_purchase", 0).fillna(0) - m.get(f"{c}_ims", 0).fillna(0)

    pdate = pd.to_datetime(m.get("document_date_purchase"), errors="coerce")
    idate = pd.to_datetime(m.get("document_date_ims"), errors="coerce")
    m["date_diff_days"] = (pdate - idate).dt.days.abs().fillna(0).astype("Int64")

    m["recon_key"] = m.apply(lambda r: make_recon_key(r.get("supplier_gstin"), r.get("document_norm")), axis=1)
    m["mismatch_type"] = "Review"
    m["match_level"] = "L0 Not matched"
    m = m.apply(lambda r: enhance_recon_row(r, amount_tol, date_tol), axis=1)

    # Mark likely matches where invoice number differs but GSTIN/date/value are close.
    m = add_probable_match_flags(m, p, i, amount_tol, date_tol)

    # Data-quality and duplicate flags.
    purchase_dups = set()
    ims_dups = set()
    if {"supplier_gstin", "document_norm"}.issubset(purchase.columns):
        purchase_dups = set(purchase[purchase.duplicated(["supplier_gstin", "document_norm"], keep=False)].apply(lambda r: make_recon_key(r.get("supplier_gstin"), r.get("document_norm")), axis=1))
    if {"supplier_gstin", "document_norm"}.issubset(ims_work.columns):
        ims_dups = set(ims_work[ims_work.duplicated(["supplier_gstin", "document_norm"], keep=False)].apply(lambda r: make_recon_key(r.get("supplier_gstin"), r.get("document_norm")), axis=1))
    m["duplicate_flag"] = m["recon_key"].apply(lambda k: "Purchase Duplicate" if k in purchase_dups else ("IMS Duplicate" if k in ims_dups else ""))
    m.loc[m["duplicate_flag"].ne("") & m["mismatch_type"].eq("Matched"), "mismatch_type"] = "Duplicate Review"

    # Presentation columns
    m["supplier_name"] = m.get("supplier_name_purchase").fillna(m.get("supplier_name_ims"))
    m["document_type"] = m.get("document_type_purchase").fillna(m.get("document_type_ims"))
    m["document_no"] = m.get("document_no_purchase").fillna(m.get("document_no_ims"))
    m["document_date"] = pd.to_datetime(m.get("document_date_purchase"), errors="coerce").fillna(pd.to_datetime(m.get("document_date_ims"), errors="coerce"))

    m["risk_score"] = m.apply(risk_score, axis=1)
    m["risk_level"] = m["risk_score"].map(risk_level)
    m["recommended_action"] = m.apply(recommend_action, axis=1)
    m["reason"] = m.apply(recommend_reason, axis=1)
    m.loc[m["mismatch_type"].eq("Probable Match - Invoice No Difference"), "recommended_action"] = "Pending"
    m.loc[m["mismatch_type"].eq("Probable Match - Invoice No Difference"), "reason"] = "Probable match by GSTIN/date/value. Keep Pending until invoice number is confirmed."
    m.loc[m["mismatch_type"].eq("Duplicate Review"), "recommended_action"] = "Pending"
    m.loc[m["mismatch_type"].eq("Duplicate Review"), "reason"] = "Duplicate document key detected. Review before final IMS action."
    m["vendor_followup_required"] = m["mismatch_type"].isin(["Only in Purchase Register", "Value / Tax Mismatch", "Tax Head Mismatch", "Value and Date Mismatch", "Only in IMS", "Probable Match - Invoice No Difference", "Duplicate Review"])
    m["final_user_action"] = m["mismatch_type"].apply(lambda x: "Accepted" if x == "Matched" else "Pending")
    m["user_remarks"] = ""
    m["confidence_score"] = m.apply(confidence_score, axis=1)
    m.loc[m["mismatch_type"].eq("Probable Match - Invoice No Difference"), "confidence_score"] = 78
    m.loc[m["mismatch_type"].eq("Duplicate Review"), "confidence_score"] = 45
    m["json_action_code"] = m["final_user_action"].apply(action_label_to_gst_code)

    priority_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    m["_risk_order"] = m["risk_level"].map(priority_order).fillna(9)
    m = m.sort_values(["_risk_order", "mismatch_type", "supplier_gstin", "document_norm"]).drop(columns=["_risk_order"])
    return m.reset_index(drop=True)


def risk_score(row) -> int:
    score = 0
    typ = str(row.get("mismatch_type", ""))
    total_tax_diff = abs(float(row.get("total_tax_diff", 0) or 0))
    taxable_diff = abs(float(row.get("taxable_value_diff", 0) or 0))
    if typ == "Only in IMS":
        score += 30
    if typ == "Only in Purchase Register":
        score += 25
    if "Value" in typ:
        score += 25
    if "Tax Head" in typ:
        score += 20
    if "Date" in typ:
        score += 10
    if total_tax_diff >= 100000 or taxable_diff >= 500000:
        score += 25
    if "CN" in str(row.get("ims_sheet_ims", "")).upper() or "credit" in str(row.get("document_type_ims", "")).lower():
        score += 15
    return min(100, score)


def risk_level(score) -> str:
    score = int(score or 0)
    if score >= 75:
        return "Critical"
    if score >= 50:
        return "High"
    if score >= 21:
        return "Medium"
    return "Low"


def recommend_action(row) -> str:
    typ = str(row.get("mismatch_type", ""))
    itc = str(row.get("itc_available_purchase", "Yes")).lower()
    if "no" in itc or "ineligible" in itc or "not" in itc:
        return "Rejected"
    if typ == "Matched":
        return "Accepted"
    if typ == "Only in IMS":
        return "Pending"
    if typ == "Only in Purchase Register":
        return "No Action"
    if typ in ["Value / Tax Mismatch", "Tax Head Mismatch", "Value and Date Mismatch", "Date Mismatch"]:
        return "Pending"
    return "Review"


def recommend_reason(row) -> str:
    typ = str(row.get("mismatch_type", ""))
    if typ == "Matched":
        return "Purchase Register and IMS values are matched within selected tolerance."
    if typ == "Only in IMS":
        return "Document appears in IMS but is not found in Purchase Register. Keep Pending until booking/vendor confirmation."
    if typ == "Only in Purchase Register":
        return "Document exists in books but not in IMS. No IMS action possible; follow up with vendor if ITC expected."
    if typ == "Tax Head Mismatch":
        return "Total tax may be close but IGST/CGST/SGST/Cess split differs. Check POS and tax head classification."
    if typ == "Value / Tax Mismatch":
        return "Amount difference detected. Compare invoice copy, credit/debit note treatment and amendment."
    if typ == "Date Mismatch":
        return "Invoice matched by GSTIN and document number, but document date differs beyond tolerance."
    if typ == "Value and Date Mismatch":
        return "Both amount and date differences detected. Manual review required before IMS action."
    return "Review required."


def confidence_score(row) -> int:
    typ = str(row.get("mismatch_type", ""))
    if typ == "Matched":
        return 100
    if typ == "Date Mismatch":
        return 82
    if typ == "Tax Head Mismatch":
        return 75
    if typ == "Value / Tax Mismatch":
        return 65
    if typ.startswith("Only"):
        return 35
    return 50


def recon_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    return df.groupby(["mismatch_type", "risk_level", "recommended_action"], dropna=False).agg(
        Records=("mismatch_type", "size"),
        Taxable_Diff=("taxable_value_diff", "sum"),
        Tax_Diff=("total_tax_diff", "sum"),
        Avg_Confidence=("confidence_score", "mean"),
    ).reset_index().round(2)


def to_excel_bytes(sheets: Dict[str, pd.DataFrame]) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        meta = pd.DataFrame([
            {"Field": "Application", "Value": APP_TITLE},
            {"Field": "Owner", "Value": COPYRIGHT_OWNER},
            {"Field": "Engine Version", "Value": ENGINE_VERSION},
            {"Field": "Generated On", "Value": datetime.now().strftime("%d-%b-%Y %H:%M:%S")},
            {"Field": "Client", "Value": st.session_state.get("client_name", "")},
            {"Field": "GSTIN", "Value": st.session_state.get("client_gstin", "")},
            {"Field": "Return Period", "Value": st.session_state.get("return_period", "")},
            {"Field": "Generated By", "Value": st.session_state.get("username", "")},
        ])
        meta.to_excel(writer, index=False, sheet_name="BAJRABHANU")
        used = {"BAJRABHANU"}
        for sheet_name, df in sheets.items():
            safe = re.sub(r"[\[\]:*?/\\]", "_", str(sheet_name))[:31] or "Sheet"
            base = safe
            i = 1
            while safe in used:
                safe = f"{base[:27]}_{i}"
                i += 1
            used.add(safe)
            data = df if isinstance(df, pd.DataFrame) and not df.empty else pd.DataFrame({"Message": ["No data available"]})
            data.to_excel(writer, index=False, sheet_name=safe)
            ws = writer.sheets[safe]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for col_cells in ws.columns:
                max_len = 0
                letter = col_cells[0].column_letter
                for cell in col_cells:
                    max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                ws.column_dimensions[letter].width = min(max_len + 2, 38)
    return buffer.getvalue()


def safe_display_df(df: pd.DataFrame, limit: int = 1000) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.head(limit).copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = out[col].dt.strftime("%d-%b-%Y").fillna("")
    return out


# =========================================================
# UI HELPERS
# =========================================================

def top_header():
    client_name = st.session_state.get("client_name", "") or "Not set"
    client_gstin = st.session_state.get("client_gstin", "") or "GSTIN pending"
    st.markdown(f"""
    <div class='gst-shell'>
        <div class='gst-top-strip'>
            <span>IMS Recon Pro</span>
            <span>•</span>
            <span>Smart GST IMS Workflow</span>
            <span>•</span>
            <span>Designed for India</span>
        </div>
        <div class='gst-masthead'>
            <div class='gst-brand'>
                <div class='gst-emblem'>🧾</div>
                <div>
                    <div class='gst-title'>Goods and Services Tax</div>
                    <div class='gst-subtitle'>IMS Recon Pro — Reconciliation, Action & GST JSON Platform</div>
                    <div class='header-note'>Premium compliance workspace for Purchase Register vs IMS JSON review</div>
                </div>
            </div>
            <div class='gst-action-wrap'>
                <div class='gst-floating-flag' title='India'></div>
            </div>
        </div>
        <div class='gst-meta-row'>
            <div class='gst-meta-card'>
                <div class='gst-meta-icon'>🗓️</div>
                <div><div class='gst-meta-label'>Today</div><div class='gst-meta-value'>{datetime.today().strftime("%d %b %Y")} • {datetime.today().strftime("%A")}</div></div>
            </div>
            <div class='gst-meta-card'>
                <div class='gst-meta-icon'>👤</div>
                <div><div class='gst-meta-label'>Logged in user</div><div class='gst-meta-value'>{st.session_state.get("display_name", "User")} • {st.session_state.get("role", "")}</div></div>
            </div>
            <div class='gst-meta-card'>
                <div class='gst-meta-icon'>🏢</div>
                <div><div class='gst-meta-label'>Client / GSTIN</div><div class='gst-meta-value'>{client_name} • {client_gstin}</div></div>
            </div>
            <div class='gst-meta-card'>
                <div class='gst-meta-icon'>🛡️</div>
                <div><div class='gst-meta-label'>System</div><div class='gst-meta-value'>GST JSON Logic Protected</div></div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def horizontal_nav():
    pages = [
        ("🏠", "Dashboard"),
        ("🔐", "Client Setup"),
        ("📤", "Upload Center"),
        ("🧾", "IMS Data Viewer"),
        ("🔄", "Reconciliation Workspace"),
        ("✅", "Action Center"),
        ("⚠️", "Risk Center"),
        ("📨", "Vendor Follow-up"),
        ("📊", "Reports & Export"),
        ("🧠", "AI Insight Desk"),
        ("👑", "Admin Panel"),
    ]
    st.markdown("<div class='gst-nav-title'>GST IMS Services</div>", unsafe_allow_html=True)
    st.markdown("<div class='gst-nav-panel'>", unsafe_allow_html=True)
    row1 = pages[:6]
    row2 = pages[6:]
    cols = st.columns(len(row1))
    for col, (icon, page) in zip(cols, row1):
        with col:
            label = f"{icon} {page}"
            if st.button(label, key=f"nav_top_{page}", use_container_width=True):
                st.session_state.page = page
                st.rerun()
    cols = st.columns(len(row2))
    for col, (icon, page) in zip(cols, row2):
        with col:
            label = f"{icon} {page}"
            if st.button(label, key=f"nav_top_{page}", use_container_width=True):
                st.session_state.page = page
                st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)


def hero_dashboard():
    st.markdown("<div class='main-shell'><div class='content-pad'><div class='watermark'>◉</div>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([3.6, 1.7, 1.5])
    with col1:
        st.markdown(f"""
        <div class='headline'>☀️ Namaste, {st.session_state.get("display_name", "User")}! 🙏</div>
        <div class='main-title'>Reconcile Today. Stay Compliant.<br>Drive Confidence.</div>
        <div class='subcopy'>AI-powered IMS reconciliation with accuracy,<br>automation & actionable insights.</div>
        """, unsafe_allow_html=True)

        b1, b2 = st.columns([1, 1])
        with b1:
            if st.button("🚀 Go to Workspace", use_container_width=True, key="hero_go_workspace"):
                st.session_state.page = "Reconciliation Workspace"
                st.rerun()
        with b2:
            if st.button("📤 Upload IMS Data", use_container_width=True, key="hero_upload_ims"):
                st.session_state.page = "Upload Center"
                st.rerun()

    with col2:
        st.markdown("<div class='shield-center'>🛡️</div>", unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <div class='feature-card'><div class='feature-title'>🧠 Smart Reconciliation</div><div class='feature-desc'>AI-driven matching with high accuracy</div></div>
        <div class='feature-card blue'><div class='feature-title'>🛡️ Risk Detection</div><div class='feature-desc'>Identify mismatches & compliance risks</div></div>
        <div class='feature-card green'><div class='feature-title'>📈 Actionable Insights</div><div class='feature-desc'>Real-time dashboards for better decisions</div></div>
        """, unsafe_allow_html=True)
    st.markdown("</div></div>", unsafe_allow_html=True)


def metric_card(icon, label, value, delta="", bg="#edf4ff", fg="#4d8df7", red=False):
    st.markdown(f"""
    <div class='metric-card'>
        <div class='metric-top'>
            <div class='metric-icon' style='background:{bg};color:{fg};'>{icon}</div>
            <div>
                <div class='metric-label'>{label}</div>
                <div class='metric-value'>{value}</div>
                <div class='metric-delta {"red" if red else ""}'>{delta}</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def page_title(title: str, subtitle: str):
    st.markdown(
        f"""
        <div style="
            background:linear-gradient(135deg,#ffffff,#f3f8ff);
            border:1px solid #cbdced;
            border-radius:22px;
            padding:20px 24px;
            margin:8px 0 18px 0;
            box-shadow:0 12px 28px rgba(7,26,61,0.09);
        ">
            <div class='section-title' style='margin:0;'>{title}</div>
            <div class='section-sub' style='margin:6px 0 0 0;'>{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def show_df(df: pd.DataFrame, limit=1000):
    if df is None or df.empty:
        st.info("No data available.")
    else:
        if len(df) > limit:
            st.caption(f"Showing first {limit:,} rows out of {len(df):,}. Use export for full data.")
        st.dataframe(safe_display_df(df, limit), use_container_width=True, hide_index=True)


# =========================================================
# LOGIN
# =========================================================

def login_page():
    st.markdown("""
    <div class='login-bg'>
        <div class='login-card'>
            <div style='text-align:center;font-size:54px;'>🇮🇳</div>
            <div class='login-title'>IMS Recon Pro</div>
            <div class='login-sub'>Secure GST IMS Reconciliation Platform<br>© @BAJRABHANU</div>
    """, unsafe_allow_html=True)

    username = st.text_input("User ID", placeholder="MainAdmin / User1 / User2")
    password = st.text_input("Password", type="password", placeholder="Enter password")

    if st.button("🔐 Login Securely", use_container_width=True):
        user = USER_MASTER.get(username)
        if user and password == user["password"]:
            st.session_state.logged_in = True
            st.session_state.username = username
            st.session_state.role = user["role"]
            st.session_state.display_name = user["name"]
            log_event("Login", "Successful login")
            load_user_state()
            st.rerun()
        else:
            st.error("Invalid User ID or Password. Please check case-sensitive credentials.")

    st.markdown("""
        </div>
    </div>
    """, unsafe_allow_html=True)


# =========================================================
# PAGES
# =========================================================

def dashboard_page():
    v10_command_center()
    v10_help_tooltips()
    hero_dashboard()
    v9_saleable_kpis()
    v9_home_modules()
    v9_json_readiness_panel()

    p, ims, recon = st.session_state.purchase_df, st.session_state.ims_df, st.session_state.recon_df
    total_itc = float(ims["total_tax"].sum()) if not ims.empty else 0
    matched = int((recon["mismatch_type"] == "Matched").sum()) if not recon.empty else 0
    pending = int((recon["recommended_action"] == "Pending").sum()) if not recon.empty else 0
    accepted = int((recon["recommended_action"] == "Accepted").sum()) if not recon.empty else 0
    highrisk = int(recon["risk_level"].isin(["High", "Critical"]).sum()) if not recon.empty else 0

    cols = st.columns(5)
    with cols[0]: metric_card("📚", "Purchase Rows", f"{len(p):,}", "Books data", "#ffefe2", "#ec8b24")
    with cols[1]: metric_card("📥", "IMS Rows", f"{len(ims):,}", st.session_state.get("ims_source",""), "#ecfaef", "#27a857")
    with cols[2]: metric_card("✅", "Matched", f"{matched:,}", "Accepted ready", "#edf4ff", "#4d8df7")
    with cols[3]: metric_card("📌", "Pending", f"{pending:,}", "Needs action", "#f4eefe", "#8b6cf7")
    with cols[4]: metric_card("⚠️", "High Risk", f"{highrisk:,}", "Review required", "#fff0ed", "#e1563a", True)

    c1, c2, c3 = st.columns([2, 2, 1.3])
    with c1:
        st.markdown("<div class='panel'><div class='panel-title'>Reconciliation Summary</div>", unsafe_allow_html=True)
        if recon.empty:
            st.info("Upload data and start reconciliation.")
        else:
            show_df(recon_summary(recon), 20)
        st.markdown("</div>", unsafe_allow_html=True)

    with c2:
        st.markdown("<div class='panel'><div class='panel-title'>Top Mismatch Reasons</div>", unsafe_allow_html=True)
        if recon.empty:
            st.info("No reconciliation data.")
        else:
            summary = recon["mismatch_type"].value_counts().reset_index()
            summary.columns = ["Mismatch Type", "Count"]
            show_df(summary, 20)
        st.markdown("</div>", unsafe_allow_html=True)

    with c3:
        st.markdown("<div class='small-card'><div class='panel-title'>Compliance Health</div>", unsafe_allow_html=True)
        health = 0 if recon.empty else round((matched / max(len(recon), 1)) * 100, 2)
        st.markdown(f"<div style='font-size:48px;font-weight:900;color:#112244;text-align:center;margin:20px 0;'>{health}%</div>", unsafe_allow_html=True)
        st.progress(int(min(100, health)))
        st.caption("Based on matched records against total reconciliation records.")
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<div class='small-card'><div class='panel-title'>AI Insight</div><br>", unsafe_allow_html=True)
        if recon.empty:
            st.write("Upload data to generate AI-like insights.")
        else:
            st.write(generate_ai_insight())
        st.markdown("</div>", unsafe_allow_html=True)


def client_setup_page():
    page_title("Client Setup", "Set client GSTIN, return period and review controls.")
    with st.form("client_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            client_name = st.text_input("Client Name", st.session_state.client_name)
        with c2:
            client_gstin = st.text_input("Client GSTIN", st.session_state.client_gstin).upper()
        with c3:
            return_period = st.text_input("Return Period", st.session_state.return_period)

        submitted = st.form_submit_button("💾 Save Client Setup", use_container_width=True)
        if submitted:
            st.session_state.client_name = client_name
            st.session_state.client_gstin = normalize_gstin(client_gstin)
            st.session_state.return_period = return_period
            save_user_state(["client_name", "client_gstin", "return_period"])
            log_event("Client Setup", "Client details saved")
            st.success("Client setup saved.")

    if st.session_state.client_gstin:
        if validate_gstin(st.session_state.client_gstin):
            st.success("GSTIN format is valid.")
        else:
            st.warning("GSTIN format appears invalid.")


def upload_center_page():
    v10_quality_dashboard()
    v10_help_tooltips()
    v9_help_box('Upload Guidance', 'Upload Purchase Register and GST IMS JSON. Review quality checks before reconciliation to avoid wrong action selection.')
    page_title("Upload Center", "Upload Purchase Register and GST IMS JSON only. The GST utility is now built inside this app.")

    st.markdown("### Step 1 — Upload source files")
    c1, c2 = st.columns(2)

    with c1:
        st.markdown("<div class='panel'><div class='panel-title'>📚 Purchase Register</div>", unsafe_allow_html=True)
        file = st.file_uploader("Upload Purchase Register", type=["xlsx", "xls", "csv"], key="purchase_upload")
        if file and st.button("Process Purchase Register", use_container_width=True):
            try:
                df = read_purchase_file(file)
                st.session_state.purchase_df = df
                st.session_state.recon_df = pd.DataFrame()
                st.session_state.action_df = pd.DataFrame()
                st.session_state.final_json_bytes = b""
                st.session_state.final_json_summary = pd.DataFrame()
                save_user_state(["purchase_df", "recon_df", "action_df", "final_json_bytes", "final_json_summary"])
                log_event("Upload", f"Purchase Register uploaded: {len(df):,} rows")
                st.success(f"Purchase Register processed: {len(df):,} rows.")
            except Exception as e:
                st.error(f"Unable to process Purchase Register: {e}")
        st.markdown("</div>", unsafe_allow_html=True)

    with c2:
        st.markdown("<div class='panel'><div class='panel-title'>🧬 GST IMS JSON</div>", unsafe_allow_html=True)
        json_file = st.file_uploader("Upload IMS JSON downloaded from GST Portal", type=["json"], key="ims_json_upload")
        if json_file and st.button("Process IMS JSON", use_container_width=True):
            try:
                raw_bytes = json_file.getvalue()
                df, records, data = parse_ims_json_bytes(raw_bytes)
                st.session_state.ims_df = df
                st.session_state.ims_source = "IMS JSON"
                st.session_state.ims_json_records = records
                st.session_state.ims_json_data = data
                st.session_state.ims_json_bytes = raw_bytes
                st.session_state.final_json_bytes = b""
                st.session_state.final_json_summary = pd.DataFrame()
                st.session_state.recon_df = pd.DataFrame()
                st.session_state.action_df = pd.DataFrame()
                if isinstance(data, dict) and data.get("rtin"):
                    st.session_state.client_gstin = normalize_gstin(data.get("rtin"))
                save_user_state(["ims_df", "ims_source", "ims_json_records", "ims_json_data", "ims_json_bytes", "final_json_bytes", "final_json_summary", "client_gstin", "recon_df", "action_df"])
                log_event("Upload", f"IMS JSON uploaded: {len(df):,} rows")
                st.success(f"IMS JSON processed: {len(df):,} rows.")
            except Exception as e:
                st.error(f"Unable to process IMS JSON: {e}")
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### Step 2 — Reconcile inside the app")
    st.info("Final process: Purchase Register + IMS JSON → in-app reconciliation → in-app action/remarks → final GST upload JSON. No .xlsm utility is required.")

    c4, c5, c6 = st.columns(3)
    with c4:
        st.session_state.amount_tolerance = st.number_input("Amount tolerance ₹", min_value=0.0, value=float(st.session_state.amount_tolerance), step=1.0, key="upload_amount_tol")
    with c5:
        st.session_state.date_tolerance = st.number_input("Date tolerance days", min_value=0, value=int(st.session_state.date_tolerance), step=1, key="upload_date_tol")
    with c6:
        st.session_state.include_amendments = st.checkbox("Include amendment records", value=bool(st.session_state.include_amendments), key="upload_include_amend")

    ready_reco = not st.session_state.purchase_df.empty and not st.session_state.ims_df.empty
    if st.button("🚀 Run Reconciliation from JSON", type="primary", use_container_width=True, disabled=not ready_reco):
        with st.spinner("Reconciling Purchase Register with IMS JSON..."):
            recon = calculate_recon(
                st.session_state.purchase_df,
                st.session_state.ims_df,
                st.session_state.amount_tolerance,
                st.session_state.date_tolerance,
                st.session_state.include_amendments,
            )
            # As per final business rule:
            # Matched = Accepted, Unmatched/Mismatch = Pending. User may manually change later.
            recon["final_user_action"] = recon["mismatch_type"].apply(lambda x: "Accepted" if x == "Matched" else "Pending")
            recon["json_action_code"] = recon["final_user_action"].apply(action_label_to_gst_code)
            st.session_state.recon_df = recon
            st.session_state.action_df = recon.copy()
            st.session_state.final_json_bytes = b""
            st.session_state.final_json_summary = pd.DataFrame()
            save_user_state(["recon_df", "action_df", "final_json_bytes", "final_json_summary"])
            log_event("Reconciliation", f"JSON reconciliation completed: {len(recon):,} rows")
        st.success(f"Reconciliation completed: {len(st.session_state.recon_df):,} rows. Now go to Action Center to review/edit actions and remarks.")

    st.markdown("---")
    page_title("Data Health Check", "Upload status and quality summary.")
    h1, h2, h3, h4 = st.columns(4)
    p = st.session_state.purchase_df
    ims = st.session_state.ims_df
    with h1: metric_card("📚", "Purchase Rows", f"{len(p):,}", "", "#ffefe2", "#ec8b24")
    with h2: metric_card("📥", "IMS JSON Rows", f"{len(ims):,}", st.session_state.ims_source, "#ecfaef", "#27a857")
    with h3:
        invalid = int((~p["gstin_valid"]).sum()) if not p.empty and "gstin_valid" in p else 0
        metric_card("⚠️", "Purchase Invalid GSTIN", f"{invalid:,}", "", "#fff0ed", "#e1563a", True)
    with h4:
        invalid = int((~ims["gstin_valid"]).sum()) if not ims.empty and "gstin_valid" in ims else 0
        metric_card("🛡️", "IMS Invalid GSTIN", f"{invalid:,}", "", "#edf4ff", "#4d8df7")

    tabs = st.tabs(["Purchase Preview", "IMS JSON Preview", "Reconciliation Preview"])
    with tabs[0]:
        show_df(st.session_state.purchase_df.head(100))
    with tabs[1]:
        show_df(st.session_state.ims_df.head(100))
    with tabs[2]:
        show_df(st.session_state.recon_df.head(100))

    st.markdown("### V7.1 Upload Validation — Taxable Value and Tax Head Wise")
    vtab1, vtab2, vtab3 = st.tabs(["Purchase Quality", "IMS Quality", "Duplicate Report"])
    with vtab1:
        show_df(upload_quality_summary(st.session_state.purchase_df, "Purchase Register"), 50)
    with vtab2:
        show_df(upload_quality_summary(st.session_state.ims_df, "IMS JSON"), 50)
        if st.session_state.ims_json_data:
            st.markdown("**IMS JSON section count**")
            show_df(ims_json_section_counts(st.session_state.ims_json_data), 50)
    with vtab3:
        dup = pd.concat([duplicate_report(st.session_state.purchase_df, "Purchase Register"), duplicate_report(st.session_state.ims_df, "IMS JSON")], ignore_index=True)
        show_df(dup, 500)

def ims_data_viewer_page():
    page_title("IMS Data Viewer", "Review uploaded and standardized data before reconciliation.")
    tabs = st.tabs(["Purchase Register", "IMS Combined", "IMS Sheet Summary"])
    with tabs[0]:
        show_df(st.session_state.purchase_df)
    with tabs[1]:
        show_df(st.session_state.ims_df)
    with tabs[2]:
        ims = st.session_state.ims_df
        if ims.empty:
            st.info("No IMS data uploaded.")
        else:
            show_df(ims.groupby(["ims_sheet", "document_type"], dropna=False).agg(
                Records=("document_no", "size"),
                Taxable=("taxable_value", "sum"),
                Tax=("total_tax", "sum"),
            ).reset_index().round(2))


def reconciliation_page():
    v10_reco_control_room()
    v9_help_box('Reconciliation Control Room', 'Run reconciliation only after both Purchase Register and IMS JSON are uploaded and validated. Review mismatch categories carefully.')
    page_title("Reconciliation Workspace", "Run IMS reconciliation only when you click Start Reconciliation.")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.session_state.amount_tolerance = st.number_input("Amount Tolerance ₹", min_value=0.0, value=float(st.session_state.amount_tolerance), step=1.0)
    with c2:
        st.session_state.date_tolerance = st.number_input("Date Tolerance Days", min_value=0, value=int(st.session_state.date_tolerance), step=1)
    with c3:
        st.session_state.include_amendments = st.checkbox("Include Amendment Sheets", value=bool(st.session_state.include_amendments))
    with c4:
        st.session_state.use_fuzzy = st.checkbox("Fuzzy Matching Later", value=False, disabled=True)

    ready = not st.session_state.purchase_df.empty and not st.session_state.ims_df.empty
    if not ready:
        st.warning("Upload Purchase Register and IMS data first.")

    if st.button("🚀 Start IMS Reconciliation", type="primary", use_container_width=True, disabled=not ready):
        with st.spinner("Running IMS reconciliation..."):
            recon = calculate_recon(
                st.session_state.purchase_df,
                st.session_state.ims_df,
                st.session_state.amount_tolerance,
                st.session_state.date_tolerance,
                st.session_state.include_amendments,
            )
            st.session_state.recon_df = recon
            st.session_state.action_df = recon.copy()
            save_user_state(["recon_df", "action_df"])
            log_event("Reconciliation", f"Reconciliation completed: {len(recon):,} rows")
        st.success(f"Reconciliation completed: {len(st.session_state.recon_df):,} rows.")

    recon = st.session_state.recon_df
    if not recon.empty:
        st.markdown("---")
        tabs = st.tabs(["All Results", "Matched", "Value Mismatch", "Only in IMS", "Only in Purchase", "High Risk"])
        with tabs[0]: show_df(recon)
        with tabs[1]: show_df(recon[recon["mismatch_type"] == "Matched"])
        with tabs[2]: show_df(recon[recon["mismatch_type"].isin(["Value / Tax Mismatch", "Tax Head Mismatch", "Value and Date Mismatch"])])
        with tabs[3]: show_df(recon[recon["mismatch_type"] == "Only in IMS"])
        with tabs[4]: show_df(recon[recon["mismatch_type"] == "Only in Purchase Register"])
        with tabs[5]: show_df(recon[recon["risk_level"].isin(["High", "Critical"])])


def action_center_page():
    v10_action_header()
    v9_help_box('Action Center Guidance', 'Use filters, bulk actions and remarks to finalize invoice-wise IMS action before generating GST upload JSON.')
    page_title("IMS Action Center", "Filter, bulk-update and finalize invoice-wise action/remarks before GST JSON generation.")
    df = st.session_state.action_df
    if df.empty:
        st.info("Run reconciliation first.")
        return

    # Ensure required columns exist.
    df = df.copy()
    if "final_user_action" not in df.columns:
        df["final_user_action"] = df.get("recommended_action", "Pending")
    if "user_remarks" not in df.columns:
        df["user_remarks"] = ""
    if "json_action_code" not in df.columns:
        df["json_action_code"] = df["final_user_action"].apply(action_label_to_gst_code)

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1: metric_card("✅", "Accepted", f"{(df['final_user_action']=='Accepted').sum():,}", "JSON A", "#ecfaef", "#27a857")
    with c2: metric_card("📌", "Pending", f"{(df['final_user_action']=='Pending').sum():,}", "JSON P", "#fff7ed", "#f4a62a")
    with c3: metric_card("🚫", "Rejected", f"{(df['final_user_action']=='Rejected').sum():,}", "JSON R", "#fff0ed", "#e1563a", True)
    with c4: metric_card("🕘", "No Action", f"{(df['final_user_action']=='No Action').sum():,}", "converted safely", "#edf4ff", "#4d8df7")
    with c5: metric_card("⚠️", "High Risk", f"{df['risk_level'].isin(['High','Critical']).sum():,}" if 'risk_level' in df else "0", "review first", "#f4eefe", "#8b6cf7")

    st.markdown("### Filters")
    f1, f2, f3, f4 = st.columns(4)
    with f1:
        action_filter = st.selectbox("Final Action", ["All"] + ACTION_VALUES, key="action_filter_v7")
    with f2:
        mismatch_options = ["All"] + sorted(df.get("mismatch_type", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
        mismatch_filter = st.selectbox("Mismatch Type", mismatch_options, key="mismatch_filter_v7")
    with f3:
        risk_options = ["All"] + sorted(df.get("risk_level", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
        risk_filter = st.selectbox("Risk Level", risk_options, key="risk_filter_v7")
    with f4:
        search_text = st.text_input("Search GSTIN / Invoice / Vendor", key="action_search_v7")

    view = df.copy()
    if action_filter != "All":
        view = view[view["final_user_action"].astype(str).eq(action_filter)]
    if mismatch_filter != "All" and "mismatch_type" in view:
        view = view[view["mismatch_type"].astype(str).eq(mismatch_filter)]
    if risk_filter != "All" and "risk_level" in view:
        view = view[view["risk_level"].astype(str).eq(risk_filter)]
    if search_text:
        stext = search_text.lower().strip()
        combined = view[[c for c in ["supplier_gstin", "supplier_name", "document_no", "document_norm"] if c in view.columns]].astype(str).agg(" ".join, axis=1).str.lower()
        view = view[combined.str.contains(re.escape(stext), na=False)]

    st.caption(f"Showing {len(view):,} rows out of {len(df):,}. Tick Select for bulk action, or directly edit Final User Action / Remarks.")

    view_cols = [
        "supplier_gstin", "supplier_name", "document_type", "document_no", "document_date",
        "taxable_value_ims", "total_tax_ims", "taxable_value_diff", "total_tax_diff",
        "mismatch_type", "match_level", "risk_level", "confidence_score", "recommended_action",
        "final_user_action", "json_action_code", "reason", "user_remarks"
    ]
    exist_cols = [c for c in view_cols if c in view.columns]
    edit_df = view[exist_cols].copy()
    edit_df.insert(0, "_row_id", edit_df.index)
    edit_df.insert(0, "Select", False)

    edited = st.data_editor(
        edit_df,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config={
            "Select": st.column_config.CheckboxColumn("Select"),
            "_row_id": st.column_config.NumberColumn("Row ID", disabled=True),
            "final_user_action": st.column_config.SelectboxColumn("Final User Action", options=ACTION_VALUES),
            "user_remarks": st.column_config.TextColumn("User Remarks"),
        },
        disabled=[c for c in edit_df.columns if c not in ["Select", "final_user_action", "user_remarks"]],
        key="action_editor_v7",
    )

    st.markdown("### Bulk Action for Selected Rows")
    b1, b2, b3 = st.columns([1, 1, 2])
    with b1:
        bulk_action = st.selectbox("Bulk final action", ACTION_VALUES, index=ACTION_VALUES.index("Pending"), key="bulk_action_v7")
    with b2:
        apply_bulk = st.button("Apply to selected", use_container_width=True)
    with b3:
        bulk_remarks = st.text_input("Optional common remarks", key="bulk_remarks_v7")

    updated = df.copy()
    # Save direct edits first.
    for _, erow in edited.iterrows():
        rid = int(erow["_row_id"])
        if rid in updated.index:
            updated.loc[rid, "final_user_action"] = erow.get("final_user_action", updated.loc[rid, "final_user_action"])
            updated.loc[rid, "user_remarks"] = erow.get("user_remarks", updated.loc[rid, "user_remarks"])

    if apply_bulk:
        selected_ids = edited.loc[edited["Select"] == True, "_row_id"].astype(int).tolist()
        if not selected_ids:
            st.warning("Please tick Select for at least one row.")
        else:
            updated.loc[selected_ids, "final_user_action"] = bulk_action
            if bulk_remarks.strip():
                updated.loc[selected_ids, "user_remarks"] = bulk_remarks.strip()
            updated["json_action_code"] = updated["final_user_action"].apply(action_label_to_gst_code)
            st.session_state.action_df = updated
            st.session_state.final_json_bytes = b""
            st.session_state.final_json_summary = pd.DataFrame()
            save_user_state(["action_df", "final_json_bytes", "final_json_summary"])
            log_event("Action Center", f"Bulk action applied to {len(selected_ids)} rows: {bulk_action}")
            st.success(f"Bulk action applied to {len(selected_ids):,} rows.")
            st.rerun()

    if st.button("💾 Save Final Actions / Remarks", type="primary", use_container_width=True):
        updated["json_action_code"] = updated["final_user_action"].apply(action_label_to_gst_code)
        st.session_state.action_df = updated
        st.session_state.final_json_bytes = b""
        st.session_state.final_json_summary = pd.DataFrame()
        save_user_state(["action_df", "final_json_bytes", "final_json_summary"])
        log_event("Action Center", "Final user actions and remarks updated")
        st.success("Final actions and remarks saved. Now go to Reports & Export for final review and GST upload JSON generation.")


def risk_center_page():
    page_title("Risk & Exception Center", "Focused review of high-risk IMS records.")
    df = st.session_state.recon_df
    if df.empty:
        st.info("Run reconciliation first.")
        return
    high = df[df["risk_level"].isin(["High", "Critical"])]
    show_df(high)
    if not high.empty:
        st.download_button(
            "Download High Risk Report",
            data=to_excel_bytes({"High Risk": high}),
            file_name="IMS_High_Risk_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def vendor_followup_page():
    v10_management_summary()
    page_title("Vendor Follow-up", "Vendor-wise pending, mismatch and follow-up list.")
    df = st.session_state.recon_df
    if df.empty:
        st.info("Run reconciliation first.")
        return
    follow = df[df["vendor_followup_required"] == True].copy()
    if follow.empty:
        st.success("No vendor follow-up cases identified.")
        return

    summary = follow.groupby(["supplier_gstin", "supplier_name"], dropna=False).agg(
        Exceptions=("mismatch_type", "size"),
        Taxable_Diff=("taxable_value_diff", "sum"),
        Tax_Diff=("total_tax_diff", "sum"),
        High_Risk=("risk_level", lambda x: int(x.isin(["High", "Critical"]).sum())),
    ).reset_index().round(2)
    show_df(summary)

    selected = st.selectbox("Generate email draft for vendor", summary["supplier_gstin"].astype(str).tolist())
    part = follow[follow["supplier_gstin"].astype(str) == str(selected)]
    if not part.empty:
        vendor_name = str(part["supplier_name"].dropna().iloc[0]) if "supplier_name" in part else ""
        email = f"""Dear {vendor_name or 'Vendor'},

During our GST IMS reconciliation for {st.session_state.get('return_period', '')}, the following documents require clarification:

Total exception cases: {len(part)}
Total tax difference / exposure: ₹{float(part['total_tax_diff'].abs().sum()):,.2f}

Request you to kindly review the invoices/credit notes/debit notes and share clarification or corrective action at the earliest.

Regards,
{st.session_state.get('display_name', '')}
"""
        st.text_area("Vendor Email Draft", email, height=220)


def reports_page():
    v10_final_json_review_ui()
    v10_management_summary()
    v9_json_readiness_panel()
    v9_report_cards()
    page_title("Reports & Final GST Upload JSON", "Final review, workpaper export and GST portal upload JSON generation.")
    p, ims, recon, action = st.session_state.purchase_df, st.session_state.ims_df, st.session_state.recon_df, st.session_state.action_df

    st.markdown("### Final Review Before GST JSON")
    if action.empty:
        st.info("No final action report available yet. Run reconciliation and save actions first.")
    else:
        review = final_json_review_table(action)
        r1, r2 = st.columns([1.2, 2])
        with r1:
            show_df(review, 20)
        with r2:
            st.info("GST JSON generation logic is the stable V6 amendment-safe logic. It remains unchanged in V7. Only your final action values are used to update GST action codes.")
            risky = action[action.get("risk_level", pd.Series(dtype=str)).isin(["High", "Critical"])] if "risk_level" in action else pd.DataFrame()
            if not risky.empty:
                st.warning(f"{len(risky):,} high/critical risk rows exist. Review them before generating JSON.")

    c1, c2 = st.columns([1, 1])
    with c1:
        st.markdown("<div class='panel'><div class='panel-title'>📊 Complete Excel Workpaper</div>", unsafe_allow_html=True)
        sheets = split_report_sheets(p, ims, recon, action)
        st.download_button(
            "📥 Download Complete IMS Workpaper",
            data=to_excel_bytes(sheets),
            file_name=f"IMS_JSON_Recon_Workpaper_V7_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)

    with c2:
        st.markdown("<div class='panel'><div class='panel-title'>🧬 Final GST Portal Upload JSON</div>", unsafe_allow_html=True)
        st.caption("Stable V6 GST schema: rtin + reqtyp SAVE + invdata. Amendment-safe fields are preserved from the original GST IMS JSON.")
        if not st.session_state.ims_json_data:
            st.warning("Please upload/process IMS JSON first.")
        elif action.empty:
            st.warning("Please run reconciliation and save final actions first.")
        else:
            source_counts = ims_json_section_counts(st.session_state.ims_json_data)
            if not source_counts.empty:
                st.markdown("**Uploaded IMS JSON section count**")
                st.dataframe(source_counts, use_container_width=True, hide_index=True)

            confirm = st.checkbox("I have reviewed final actions and want to generate GST upload JSON", key="confirm_json_v7")
            if st.button("⚙️ Generate Final GST Upload JSON", type="primary", use_container_width=True, disabled=not confirm):
                try:
                    # DO NOT CHANGE: stable V6 amendment-safe generator.
                    json_bytes, summary = generate_gst_upload_json_from_final_actions(st.session_state.ims_json_data, st.session_state.action_df)
                    st.session_state.final_json_bytes = json_bytes
                    st.session_state.final_json_summary = summary
                    save_user_state(["final_json_bytes", "final_json_summary"])
                    log_event("GST JSON", f"Final GST upload JSON generated: {len(summary):,} records")
                    st.success(f"Final GST upload JSON generated. Records included: {len(summary):,}")
                except Exception as e:
                    st.error(f"Unable to generate final JSON: {e}")

            if st.session_state.final_json_bytes:
                action_counts = generated_json_action_counts(st.session_state.final_json_bytes)
                if not action_counts.empty:
                    st.markdown("**Generated GST upload JSON action count**")
                    st.dataframe(action_counts, use_container_width=True, hide_index=True)
                st.download_button(
                    "⬇️ Download Final GST Portal Upload JSON",
                    data=st.session_state.final_json_bytes,
                    file_name=f"IMS_Final_Action_Upload_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                    mime="application/json",
                    use_container_width=True,
                )
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### Upload & Duplicate Validation")
    t1, t2, t3 = st.tabs(["Purchase Quality", "IMS Quality", "Duplicates"])
    with t1:
        show_df(upload_quality_summary(p, "Purchase Register"), 50)
    with t2:
        show_df(upload_quality_summary(ims, "IMS JSON"), 50)
    with t3:
        dup = pd.concat([duplicate_report(p, "Purchase Register"), duplicate_report(ims, "IMS JSON")], ignore_index=True)
        show_df(dup, 500)

    if isinstance(st.session_state.final_json_summary, pd.DataFrame) and not st.session_state.final_json_summary.empty:
        st.markdown("### Final JSON Update Summary")
        show_df(st.session_state.final_json_summary.groupby(["Section", "Final Action", "GST JSON Action Code"], dropna=False).size().reset_index(name="Records"), 100)


def ai_insight_page():
    v10_management_summary()
    v10_help_tooltips()
    page_title("AI Insight Desk", "Rule-based smart GST IMS insights without external API.")
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    st.write(generate_ai_insight(long=True))
    st.markdown("</div>", unsafe_allow_html=True)

    q = st.text_input("Ask IMS Insight Desk", placeholder="Example: Which invoices should I accept?")
    if q:
        ql = q.lower()
        df = st.session_state.recon_df
        if df.empty:
            st.info("Run reconciliation first.")
        elif "accept" in ql:
            show_df(df[df["recommended_action"] == "Accepted"])
        elif "pending" in ql:
            show_df(df[df["recommended_action"] == "Pending"])
        elif "reject" in ql:
            show_df(df[df["recommended_action"] == "Rejected"])
        elif "risk" in ql:
            show_df(df[df["risk_level"].isin(["High", "Critical"])])
        elif "vendor" in ql:
            vendor_followup_page()
        else:
            st.write(generate_ai_insight(long=True))


def generate_ai_insight(long: bool = False) -> str:
    df = st.session_state.recon_df
    if df.empty:
        return "No reconciliation has been run yet. Upload Purchase Register and IMS data, then click Start IMS Reconciliation."
    total = len(df)
    matched = int((df["mismatch_type"] == "Matched").sum())
    pending = int((df["recommended_action"] == "Pending").sum())
    high = int(df["risk_level"].isin(["High", "Critical"]).sum())
    only_ims = int((df["mismatch_type"] == "Only in IMS").sum())
    only_purchase = int((df["mismatch_type"] == "Only in Purchase Register").sum())
    health = round((matched / max(total, 1)) * 100, 2)
    base = (
        f"IMS Health Score is {health}%. Out of {total:,} reconciled documents, "
        f"{matched:,} are clean matches, {pending:,} should be kept pending/reviewed, "
        f"and {high:,} are high/critical risk cases."
    )
    if not long:
        return base
    return (
        base
        + f"\n\nKey observations:\n"
        + f"- {only_ims:,} documents are appearing in IMS but not in Purchase Register.\n"
        + f"- {only_purchase:,} documents are booked in Purchase Register but not found in IMS.\n"
        + "- Accept matched invoices first, keep value/tax mismatch cases pending, and send vendor follow-up for books-only cases.\n"
        + "- Review credit notes and amendment sheets separately before final upload through final GST upload JSON."
    )


def admin_page():
    page_title("Admin Panel", "Data reset, audit log and user controls.")

    st.markdown("### Audit Log")
    show_df(load_audit("" if st.session_state.role == "Main Admin" else st.session_state.username), 500)

    st.markdown("### Reset Data")
    st.warning("Reset will delete saved data for the current logged-in user.")
    confirm = st.text_input("Type DELETE to confirm reset")
    if st.button("🗑️ Reset My Complete Data", use_container_width=True):
        if confirm == "DELETE":
            username = st.session_state.username
            db_delete_user(username)
            log_event("Reset", "User data reset")
            for key in ["purchase_df", "ims_df", "recon_df", "action_df"]:
                st.session_state[key] = pd.DataFrame()
            st.session_state.ims_source = ""
            st.success("Your saved data has been deleted.")
        else:
            st.error("Please type DELETE exactly.")




# =========================================================
# V9 SALEABLE UI HELPERS — UI ONLY, NO GST JSON LOGIC CHANGE
# =========================================================

def v9_status_bool(value) -> bool:
    try:
        if isinstance(value, pd.DataFrame):
            return not value.empty
        return bool(value)
    except Exception:
        return False


def v9_workflow_tracker():
    p_ready = v9_status_bool(st.session_state.get("purchase_df", pd.DataFrame()))
    ims_ready = v9_status_bool(st.session_state.get("ims_df", pd.DataFrame()))
    recon_ready = v9_status_bool(st.session_state.get("recon_df", pd.DataFrame()))
    action_ready = v9_status_bool(st.session_state.get("action_df", pd.DataFrame()))
    client_ready = bool(st.session_state.get("client_gstin", ""))

    steps = [
        ("01", "Client Setup", client_ready, st.session_state.get("page") == "Client Setup"),
        ("02", "Upload Purchase", p_ready, st.session_state.get("page") == "Upload Center"),
        ("03", "Upload IMS JSON", ims_ready, st.session_state.get("page") == "Upload Center"),
        ("04", "Reconciliation", recon_ready, st.session_state.get("page") == "Reconciliation Workspace"),
        ("05", "Action Review", action_ready, st.session_state.get("page") == "Action Center"),
        ("06", "GST JSON", False, st.session_state.get("page") == "Reports & Export"),
    ]

    html = ["<div class='v9-workflow'><div class='v9-workflow-title'>🚀 Guided IMS Workflow</div><div class='v9-step-grid'>"]
    for num, label, done, active in steps:
        cls = "done" if done else ("active" if active else "pending")
        status = "Completed" if done else ("Action Required" if active else "Pending")
        icon = "✓" if done else num
        html.append(f"""
        <div class='v9-step {cls}'>
            <div class='v9-step-num'>{icon}</div>
            <div class='v9-step-label'>{label}</div>
            <div class='v9-step-status'>{status}</div>
        </div>
        """)
    html.append("</div></div>")
    st.markdown("".join(html), unsafe_allow_html=True)


def v9_saleable_kpis():
    p = st.session_state.get("purchase_df", pd.DataFrame())
    ims = st.session_state.get("ims_df", pd.DataFrame())
    recon = st.session_state.get("recon_df", pd.DataFrame())
    action = st.session_state.get("action_df", pd.DataFrame())

    matched = int((recon.get("mismatch_type", pd.Series(dtype=str)) == "Matched").sum()) if isinstance(recon, pd.DataFrame) and not recon.empty else 0
    pending = int((action.get("final_user_action", pd.Series(dtype=str)) == "Pending").sum()) if isinstance(action, pd.DataFrame) and not action.empty else 0
    accepted = int((action.get("final_user_action", pd.Series(dtype=str)) == "Accepted").sum()) if isinstance(action, pd.DataFrame) and not action.empty else 0
    highrisk = int(action.get("risk_level", pd.Series(dtype=str)).isin(["High", "Critical"]).sum()) if isinstance(action, pd.DataFrame) and not action.empty and "risk_level" in action else 0

    st.markdown(f"""
    <div class='v9-kpi-strip'>
        <div class='v9-kpi'><div class='v9-kpi-label'>Purchase Register</div><div class='v9-kpi-value'>{len(p):,}</div><div class='v9-kpi-note'>Books records loaded</div></div>
        <div class='v9-kpi'><div class='v9-kpi-label'>IMS JSON</div><div class='v9-kpi-value'>{len(ims):,}</div><div class='v9-kpi-note'>Portal records loaded</div></div>
        <div class='v9-kpi'><div class='v9-kpi-label'>Accepted / Matched</div><div class='v9-kpi-value'>{accepted:,}</div><div class='v9-kpi-note'>{matched:,} system matches</div></div>
        <div class='v9-kpi'><div class='v9-kpi-label'>Pending / Risk</div><div class='v9-kpi-value'>{pending:,}</div><div class='v9-kpi-note'>{highrisk:,} high-risk cases</div></div>
    </div>
    """, unsafe_allow_html=True)


def v9_home_modules():
    st.markdown("""
    <div class='v9-module-grid'>
        <div class='v9-module-card'>
            <div class='v9-module-icon'>📤</div>
            <div class='v9-module-title'>Smart Upload Center</div>
            <div class='v9-module-desc'>Upload Purchase Register and GST IMS JSON with quality checks, duplicate review and section-wise visibility.</div>
            <div class='v9-module-badge'>Upload → Validate</div>
        </div>
        <div class='v9-module-card'>
            <div class='v9-module-icon'>🔄</div>
            <div class='v9-module-title'>Reconciliation Control Room</div>
            <div class='v9-module-desc'>Review matched, pending, mismatch, duplicate and risk cases in a structured and user-friendly flow.</div>
            <div class='v9-module-badge'>Recon → Review</div>
        </div>
        <div class='v9-module-card'>
            <div class='v9-module-icon'>✅</div>
            <div class='v9-module-title'>Action Center</div>
            <div class='v9-module-desc'>Use filters, manual actions, remarks and bulk review to finalize invoice-wise IMS actions.</div>
            <div class='v9-module-badge'>Action → Finalize</div>
        </div>
        <div class='v9-module-card'>
            <div class='v9-module-icon'>⚠️</div>
            <div class='v9-module-title'>Risk Desk</div>
            <div class='v9-module-desc'>Identify value mismatches, tax-head mismatches, only-in-IMS cases and vendor follow-up requirements.</div>
            <div class='v9-module-badge'>Risk → Resolve</div>
        </div>
        <div class='v9-module-card'>
            <div class='v9-module-icon'>📊</div>
            <div class='v9-module-title'>Professional Reports</div>
            <div class='v9-module-desc'>Download Excel workpapers with summary, final action, mismatch, pending, risk and audit reports.</div>
            <div class='v9-module-badge'>Report → Export</div>
        </div>
        <div class='v9-module-card'>
            <div class='v9-module-icon'>🧾</div>
            <div class='v9-module-title'>GST JSON Output</div>
            <div class='v9-module-desc'>Generate GST portal-ready JSON after final review while keeping the confirmed JSON logic protected.</div>
            <div class='v9-module-badge'>Review → JSON</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def v9_json_readiness_panel():
    action = st.session_state.get("action_df", pd.DataFrame())
    ready = isinstance(action, pd.DataFrame) and not action.empty
    if ready:
        accepted = int((action.get("final_user_action", pd.Series(dtype=str)) == "Accepted").sum())
        pending = int((action.get("final_user_action", pd.Series(dtype=str)) == "Pending").sum())
        rejected = int((action.get("final_user_action", pd.Series(dtype=str)) == "Rejected").sum())
        status_line = f"Accepted: {accepted:,} • Pending: {pending:,} • Rejected: {rejected:,}"
    else:
        status_line = "Run reconciliation and review actions first"

    st.markdown(f"""
    <div class='v9-readiness'>
        <div class='v9-readiness-title'>🛡️ Final GST JSON Readiness</div>
        <div class='v9-check-grid'>
            <div class='v9-check'><div class='v9-check-icon'>✅</div><div class='v9-check-label'>GST upload structure protected</div></div>
            <div class='v9-check'><div class='v9-check-icon'>🧾</div><div class='v9-check-label'>rtin / reqtyp / invdata preserved</div></div>
            <div class='v9-check'><div class='v9-check-icon'>🔐</div><div class='v9-check-label'>Amendment-safe section handling</div></div>
            <div class='v9-check'><div class='v9-check-icon'>📌</div><div class='v9-check-label'>{status_line}</div></div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def v9_help_box(title, text):
    st.markdown(f"""
    <div class='v9-help-box'>
        <div class='v9-help-title'>{title}</div>
        <div class='v9-help-text'>{text}</div>
    </div>
    """, unsafe_allow_html=True)


def v9_report_cards():
    st.markdown("""
    <div class='v9-report-grid'>
        <div class='v9-report-card'><div class='v9-report-title'>📘 Final Action Report</div><div class='v9-report-desc'>Invoice-wise final action, remarks and recommended action summary.</div></div>
        <div class='v9-report-card'><div class='v9-report-title'>⚠️ Risk & Exception Report</div><div class='v9-report-desc'>High-risk cases, mismatches, duplicates and vendor follow-up items.</div></div>
        <div class='v9-report-card'><div class='v9-report-title'>🧾 JSON Upload Summary</div><div class='v9-report-desc'>Action summary and records prepared for GST portal JSON generation.</div></div>
    </div>
    """, unsafe_allow_html=True)





# =========================================================
# V10 ADVANCED SALEABLE UI HELPERS — UI ONLY
# Final GST JSON generation logic is intentionally untouched.
# =========================================================

def v10_df_len(key: str) -> int:
    try:
        df = st.session_state.get(key, pd.DataFrame())
        return len(df) if isinstance(df, pd.DataFrame) else 0
    except Exception:
        return 0


def v10_safe_sum(df: pd.DataFrame, col: str) -> float:
    try:
        if isinstance(df, pd.DataFrame) and col in df.columns:
            return float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum())
    except Exception:
        pass
    return 0.0


def v10_quality_score(df: pd.DataFrame) -> int:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return 0
    total = max(len(df), 1)
    score = 100
    for col in ["supplier_gstin", "document_no", "document_date"]:
        if col in df.columns:
            blanks = df[col].astype(str).str.strip().isin(["", "nan", "None", "NaT"]).sum()
            score -= int((blanks / total) * 18)
    if "supplier_gstin" in df.columns:
        invalid = (~df["supplier_gstin"].astype(str).str.upper().str.match(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]$", na=False)).sum()
        score -= int((invalid / total) * 20)
    if {"supplier_gstin", "document_no"}.issubset(df.columns):
        dups = df.duplicated(["supplier_gstin", "document_no"], keep=False).sum()
        score -= int((dups / total) * 15)
    return max(0, min(100, score))


def v10_score_class(score: int) -> str:
    if score >= 90:
        return ""
    if score >= 70:
        return "warn"
    return "bad"


def v10_command_center():
    st.markdown("""
    <div class='v10-command-center'>
        <div class='v10-command-title'>⚡ IMS Recon Pro Command Center</div>
        <div class='v10-command-sub'>A premium workflow built for GST IMS reconciliation, exception review, action control, reporting and final GST upload preparation.</div>
        <div class='v10-action-grid'>
            <div class='v10-action-card'><div class='v10-action-icon'>🚀</div><div class='v10-action-title'>Start New Reconciliation</div><div class='v10-action-desc'>Begin client setup, upload data and run IMS matching in a guided workflow.</div></div>
            <div class='v10-action-card'><div class='v10-action-icon'>📤</div><div class='v10-action-title'>Upload & Validate</div><div class='v10-action-desc'>Check GSTIN, invoice details, duplicates and tax values before matching.</div></div>
            <div class='v10-action-card'><div class='v10-action-icon'>✅</div><div class='v10-action-title'>Review Actions</div><div class='v10-action-desc'>Finalize Accepted, Pending and Rejected actions with remarks and filters.</div></div>
            <div class='v10-action-card'><div class='v10-action-icon'>🧾</div><div class='v10-action-title'>Generate Output</div><div class='v10-action-desc'>Prepare reports and GST portal-ready JSON after final review.</div></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if st.button("📌 Client Setup", use_container_width=True, key="v10_client_setup"):
            st.session_state.page = "Client Setup"
            st.rerun()
    with c2:
        if st.button("📤 Upload Data", use_container_width=True, key="v10_upload_data"):
            st.session_state.page = "Upload Center"
            st.rerun()
    with c3:
        if st.button("🔄 Run Reco", use_container_width=True, key="v10_run_reco"):
            st.session_state.page = "Reconciliation Workspace"
            st.rerun()
    with c4:
        if st.button("🧾 Final JSON", use_container_width=True, key="v10_final_json"):
            st.session_state.page = "Reports & Export"
            st.rerun()


def v10_quality_dashboard():
    p = st.session_state.get("purchase_df", pd.DataFrame())
    ims = st.session_state.get("ims_df", pd.DataFrame())
    p_score = v10_quality_score(p)
    ims_score = v10_quality_score(ims)

    def card(title, df, score, source):
        total_tax = sum(v10_safe_sum(df, c) for c in ["igst", "cgst", "sgst", "cess"])
        taxable = v10_safe_sum(df, "taxable_value")
        invoice_value = v10_safe_sum(df, "invoice_value")
        cls = v10_score_class(score)
        return f"""
        <div class='v10-quality-card'>
            <div class='v10-quality-head'>
                <div>
                    <div class='v10-quality-title'>{title}</div>
                    <div style='font-size:13px;color:#60748f;margin-top:5px;'>Data health, amount and tax summary</div>
                </div>
                <div class='v10-quality-score {cls}'>{score}%</div>
            </div>
            <div class='v10-mini-grid'>
                <div class='v10-mini-stat'><div class='v10-mini-label'>Records</div><div class='v10-mini-value'>{len(df):,}</div></div>
                <div class='v10-mini-stat'><div class='v10-mini-label'>Taxable</div><div class='v10-mini-value'>₹{taxable:,.0f}</div></div>
                <div class='v10-mini-stat'><div class='v10-mini-label'>Total Tax</div><div class='v10-mini-value'>₹{total_tax:,.0f}</div></div>
                <div class='v10-mini-stat'><div class='v10-mini-label'>Invoice Value</div><div class='v10-mini-value'>₹{invoice_value:,.0f}</div></div>
                <div class='v10-mini-stat'><div class='v10-mini-label'>IGST</div><div class='v10-mini-value'>₹{v10_safe_sum(df, "igst"):,.0f}</div></div>
                <div class='v10-mini-stat'><div class='v10-mini-label'>CGST + SGST</div><div class='v10-mini-value'>₹{v10_safe_sum(df, "cgst") + v10_safe_sum(df, "sgst"):,.0f}</div></div>
            </div>
        </div>
        """

    st.markdown(f"""
    <div class='v10-quality-grid'>
        {card("Purchase Register Quality", p, p_score, "Purchase")}
        {card("IMS JSON Quality", ims, ims_score, "IMS")}
    </div>
    """, unsafe_allow_html=True)


def v10_empty_state(title: str, text: str, icon: str = "📭"):
    st.markdown(f"""
    <div class='v10-empty-state'>
        <div class='v10-empty-icon'>{icon}</div>
        <div class='v10-empty-title'>{title}</div>
        <div class='v10-empty-text'>{text}</div>
    </div>
    """, unsafe_allow_html=True)


def v10_reco_control_room():
    recon = st.session_state.get("recon_df", pd.DataFrame())
    action = st.session_state.get("action_df", pd.DataFrame())

    if not isinstance(recon, pd.DataFrame) or recon.empty:
        v10_empty_state("Reconciliation not started yet", "Upload Purchase Register and IMS JSON, then run reconciliation to open the control room.", "🔄")
        return

    def count_col(df, col, value):
        try:
            return int((df[col] == value).sum()) if col in df.columns else 0
        except Exception:
            return 0

    matched = count_col(recon, "mismatch_type", "Matched")
    only_ims = count_col(recon, "mismatch_type", "Only in IMS")
    only_purchase = count_col(recon, "mismatch_type", "Only in Purchase")
    value_mismatch = int(recon.get("mismatch_type", pd.Series(dtype=str)).astype(str).str.contains("Value", case=False, na=False).sum()) if "mismatch_type" in recon.columns else 0
    highrisk = int(action.get("risk_level", pd.Series(dtype=str)).isin(["High", "Critical"]).sum()) if isinstance(action, pd.DataFrame) and not action.empty and "risk_level" in action.columns else 0

    st.markdown(f"""
    <div class='v10-control-room'>
        <div class='v10-control-main'>
            <div class='v10-control-title'>🎛️ Reconciliation Control Room</div>
            <div style='font-size:14px;color:#60748f;line-height:1.5;'>Focus on exception areas first. Use the Action Center to finalize Accepted, Pending and Rejected actions.</div>
            <div class='v10-badge-row'>
                <span class='v10-filter-badge green'>✅ Matched: {matched:,}</span>
                <span class='v10-filter-badge orange'>🟠 Only in IMS: {only_ims:,}</span>
                <span class='v10-filter-badge purple'>📘 Only in Purchase: {only_purchase:,}</span>
                <span class='v10-filter-badge red'>⚠️ Value Mismatch: {value_mismatch:,}</span>
                <span class='v10-filter-badge red'>🔥 High Risk: {highrisk:,}</span>
            </div>
        </div>
        <div class='v10-control-side'>
            <div class='v10-control-title'>🧭 Suggested Review Order</div>
            <div class='v10-badge-row'>
                <span class='v10-filter-badge red'>1. High Risk</span>
                <span class='v10-filter-badge orange'>2. Value Mismatch</span>
                <span class='v10-filter-badge purple'>3. Only in IMS</span>
                <span class='v10-filter-badge green'>4. Matched</span>
            </div>
            <div style='font-size:13px;color:#60748f;line-height:1.55;'>This review flow helps users reach final action faster and reduces manual checking effort.</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def v10_action_header():
    action = st.session_state.get("action_df", pd.DataFrame())
    if not isinstance(action, pd.DataFrame) or action.empty:
        v10_empty_state("No action table available", "Run reconciliation first, then review invoice-wise actions here.", "✅")
        return

    def c(value):
        try:
            return int((action["final_user_action"] == value).sum()) if "final_user_action" in action.columns else 0
        except Exception:
            return 0

    accepted, pending, rejected = c("Accepted"), c("Pending"), c("Rejected")
    review = c("Review")
    no_action = c("No Action")

    st.markdown(f"""
    <div class='v10-control-main' style='margin:14px 0 18px 0;'>
        <div class='v10-control-title'>✅ Action Center Command Bar</div>
        <div style='font-size:14px;color:#60748f;line-height:1.5;'>Use filters, manual action and remarks to finalize invoices before GST JSON generation.</div>
        <div class='v10-badge-row'>
            <span class='v10-filter-badge green'>Accepted: {accepted:,}</span>
            <span class='v10-filter-badge orange'>Pending: {pending:,}</span>
            <span class='v10-filter-badge red'>Rejected: {rejected:,}</span>
            <span class='v10-filter-badge purple'>Review: {review:,}</span>
            <span class='v10-filter-badge'>No Action: {no_action:,}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)


def v10_final_json_review_ui():
    action = st.session_state.get("action_df", pd.DataFrame())
    ims = st.session_state.get("ims_df", pd.DataFrame())

    accepted = pending = rejected = no_action = review = 0
    if isinstance(action, pd.DataFrame) and not action.empty and "final_user_action" in action.columns:
        accepted = int((action["final_user_action"] == "Accepted").sum())
        pending = int((action["final_user_action"] == "Pending").sum())
        rejected = int((action["final_user_action"] == "Rejected").sum())
        no_action = int((action["final_user_action"] == "No Action").sum())
        review = int((action["final_user_action"] == "Review").sum())

    sections = 0
    try:
        if isinstance(ims, pd.DataFrame) and "ims_section" in ims.columns:
            sections = ims["ims_section"].nunique()
    except Exception:
        sections = 0

    st.markdown(f"""
    <div class='v10-json-review'>
        <div class='v10-json-title'>🧾 Final GST Upload JSON Review</div>
        <div class='v10-json-checks'>
            <div class='v10-json-check'><div class='v10-json-check-icon'>✅</div><div class='v10-json-check-label'>Accepted<br>{accepted:,}</div></div>
            <div class='v10-json-check'><div class='v10-json-check-icon'>🟠</div><div class='v10-json-check-label'>Pending<br>{pending:,}</div></div>
            <div class='v10-json-check'><div class='v10-json-check-icon'>🔴</div><div class='v10-json-check-label'>Rejected<br>{rejected:,}</div></div>
            <div class='v10-json-check'><div class='v10-json-check-icon'>📦</div><div class='v10-json-check-label'>IMS Sections<br>{sections:,}</div></div>
            <div class='v10-json-check'><div class='v10-json-check-icon'>🛡️</div><div class='v10-json-check-label'>GST JSON Logic<br>Protected</div></div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def v10_management_summary():
    p = st.session_state.get("purchase_df", pd.DataFrame())
    ims = st.session_state.get("ims_df", pd.DataFrame())
    action = st.session_state.get("action_df", pd.DataFrame())

    accepted = pending = rejected = 0
    if isinstance(action, pd.DataFrame) and not action.empty and "final_user_action" in action.columns:
        accepted = int((action["final_user_action"] == "Accepted").sum())
        pending = int((action["final_user_action"] == "Pending").sum())
        rejected = int((action["final_user_action"] == "Rejected").sum())

    summary = (
        f"For the selected period, Purchase Register has {len(p):,} records and IMS JSON has {len(ims):,} records. "
        f"Based on current action review, {accepted:,} records are marked Accepted, {pending:,} records are marked Pending, "
        f"and {rejected:,} records are marked Rejected. The final GST upload JSON should be generated only after completing invoice-wise review."
    )

    st.markdown(f"""
    <div class='v10-management-summary'>
        <div class='v10-management-title'>📝 Management Summary</div>
        <div class='v10-management-text'>{summary}</div>
    </div>
    """, unsafe_allow_html=True)


def v10_help_tooltips():
    st.markdown("""
    <div class='v10-tooltip-grid'>
        <div class='v10-tooltip'><div class='v10-tooltip-title'>What is IMS JSON?</div><div class='v10-tooltip-text'>The file downloaded from GST portal containing inward supply records and action fields.</div></div>
        <div class='v10-tooltip'><div class='v10-tooltip-title'>When to mark Pending?</div><div class='v10-tooltip-text'>Use Pending for unmatched, disputed or review-required invoices.</div></div>
        <div class='v10-tooltip'><div class='v10-tooltip-title'>When to Accept?</div><div class='v10-tooltip-text'>Use Accepted when invoice details match books and credit/action is acceptable.</div></div>
        <div class='v10-tooltip'><div class='v10-tooltip-title'>Final JSON?</div><div class='v10-tooltip-text'>Generated after action review and uploaded back to GST portal.</div></div>
    </div>
    """, unsafe_allow_html=True)



# =========================================================
# MAIN
# =========================================================

def main():
    init_db()
    init_state()
    inject_css()

    if not st.session_state.logged_in:
        login_page()
        return

    top_header()
    horizontal_nav()

    page = st.session_state.page
    if page == "Dashboard":
        dashboard_page()
    elif page == "Client Setup":
        client_setup_page()
    elif page == "Upload Center":
        upload_center_page()
    elif page == "IMS Data Viewer":
        ims_data_viewer_page()
    elif page == "Reconciliation Workspace":
        reconciliation_page()
    elif page == "Action Center":
        action_center_page()
    elif page == "Risk Center":
        risk_center_page()
    elif page == "Vendor Follow-up":
        vendor_followup_page()
    elif page == "Reports & Export":
        reports_page()
    elif page == "AI Insight Desk":
        ai_insight_page()
    elif page == "Admin Panel":
        admin_page()

    st.markdown(f"""
    <div class='footer-bar'>
        <div style='display:flex;justify-content:space-around;gap:20px;flex-wrap:wrap;'>
            <div class='foot-item'><div style='font-size:26px;'>🛡️</div><div><div class='foot-main'>Secure</div><div class='foot-sub'>Enterprise-grade control</div></div></div>
            <div class='foot-item'><div style='font-size:26px;'>✅</div><div><div class='foot-main'>Compliant</div><div class='foot-sub'>GSTN workflow aligned</div></div></div>
            <div class='foot-item'><div style='font-size:26px;'>🔄</div><div><div class='foot-main'>Reliable</div><div class='foot-sub'>JSON + export ready</div></div></div>
            <div class='foot-item'><div style='font-size:26px;'>✨</div><div><div class='foot-main'>Smart</div><div class='foot-sub'>AI-like insights</div></div></div>
        </div>
    </div>
    <div style='text-align:center;color:#5d718e;font-size:14px;margin-top:14px;padding-bottom:10px;'>
        © 2026 IMS Recon Pro • {COPYRIGHT_OWNER} • Designed for India • Built for Compliance
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
